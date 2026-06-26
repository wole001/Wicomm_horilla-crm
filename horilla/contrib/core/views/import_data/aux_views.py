"""Auxiliary import views: model fields, value mapping, downloads, history, templates."""

# Standard library imports
import csv
import logging
import traceback
from io import BytesIO

# Third-party imports (Django)
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.files.storage import default_storage
from django.views.generic import TemplateView, View

# Third-party imports (other)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# First party imports (Horilla)
from horilla.apps import apps
from horilla.contrib.generics.views import HorillaListView
from horilla.db.models import CharField, ForeignKey
from horilla.registry.feature import FEATURE_REGISTRY
from horilla.shortcuts import render
from horilla.urls import reverse_lazy
from horilla.utils.decorators import (
    htmx_required,
    method_decorator,
    permission_required_or_denied,
)
from horilla.utils.translation import gettext_lazy as _
from horilla.web import HttpNotFound, HttpResponse

from ...models import ImportHistory

# Local imports
from .base import IMPORT_EXCLUDED_FIELDS

logger = logging.getLogger(__name__)


class GetModelFieldsView(LoginRequiredMixin, View):
    """HTMX view to get model fields when module is selected"""

    def get(self, request, *args, **kwargs):
        """Retrieve model fields for the selected module"""
        import_data = request.session.get("import_data", {})
        unique_values = import_data.get("unique_values", {})
        field_name = request.GET.get("field_name", "")

        # Get file header from either GET parameter or form data
        file_header = request.GET.get(f"file_header_{field_name}", "")
        if not file_header:
            file_header = request.GET.get("file_header_", "")

        app_label = request.GET.get("app_label", "")
        module = request.GET.get("module", "")

        if not all([file_header, field_name, app_label, module]):
            missing_params = [
                param
                for param, value in [
                    ("file_header", file_header),
                    ("field_name", field_name),
                    ("app_label", app_label),
                    ("module", module),
                ]
                if not value
            ]
            raise HttpNotFound(f"Missing parameters: {', '.join(missing_params)}")

        try:
            model = apps.get_model(app_label, module)
            field = next((f for f in model._meta.fields if f.name == field_name), None)
            if not field:
                return render(
                    request,
                    "common/message_fragment.html",
                    {
                        "message": f"Field not found: {field_name}",
                        "variant": "border",
                    },
                )

            unique_file_values = unique_values.get(file_header, [])
            is_choice_field = isinstance(field, CharField) and field.choices
            is_foreign_key = isinstance(field, ForeignKey)

            context = {
                "field": {
                    "name": field_name,
                    "verbose_name": field.verbose_name.title(),
                    "is_choice_field": is_choice_field,
                    "is_foreign_key": is_foreign_key,
                    "choices": (
                        [
                            {"value": value, "label": label}
                            for value, label in field.choices
                        ]
                        if is_choice_field
                        else []
                    ),
                    "foreign_key_choices": (
                        [
                            {"id": instance.pk, "display": str(instance)}
                            for instance in field.related_model.objects.all()
                        ]
                        if is_foreign_key
                        else []
                    ),
                    "unique_file_values": unique_file_values,
                },
                "choice_mappings": import_data.get("choice_mappings", {}).get(
                    field_name, {}
                ),
                "fk_mappings": import_data.get("fk_mappings", {}).get(field_name, {}),
                "auto_choice_mappings": import_data.get("auto_choice_mappings", {}).get(
                    field_name, {}
                ),
                "auto_fk_mappings": import_data.get("auto_fk_mappings", {}).get(
                    field_name, {}
                ),
            }

            return render(request, "import/value_mapping_partial.html", context)
        except Exception as e:
            tb = traceback.format_exc()
            logger.error(tb)
            return render(
                request,
                "common/message_fragment.html",
                {"message": f"Error: {e!s}", "variant": "border"},
            )


@method_decorator(
    permission_required_or_denied("core.can_view_horilla_import"),
    name="dispatch",
)
class UpdateFieldStatusView(LoginRequiredMixin, View):
    """HTMX view to update field mapping status"""

    def post(self, request, *args, **kwargs):
        """Update field mapping status based on user input"""
        field_name = request.GET.get("field_name")
        file_header = request.POST.get(f"file_header_{field_name}")

        if file_header:
            return render(
                request,
                "common/message_fragment.html",
                {"message": _("Mapped"), "variant": "success_badge"},
            )
        return render(
            request,
            "common/message_fragment.html",
            {"message": _("Not Mapped"), "variant": "badge"},
        )


@method_decorator(
    permission_required_or_denied("core.can_view_horilla_import"),
    name="dispatch",
)
class GetUniqueValuesView(LoginRequiredMixin, View):
    """HTMX view to get unique values for a selected file header with auto-mapping"""

    def get(self, request, *args, **kwargs):
        """Retrieve unique values for the selected file header"""
        import_data = request.session.get("import_data", {})
        unique_values = import_data.get("unique_values", {})
        field_name = request.GET.get("field_name", "")

        # Get file header from either GET parameter or form data
        file_header = request.GET.get(f"file_header_{field_name}", "")
        if not file_header:
            file_header = request.GET.get("file_header", "")

        app_label = request.GET.get("app_label", "")
        module = request.GET.get("module", "")

        if not all([file_header, field_name, app_label, module]):
            missing_params = [
                param
                for param, value in [
                    ("file_header", file_header),
                    ("field_name", field_name),
                    ("app_label", app_label),
                    ("module", module),
                ]
                if not value
            ]
            raise HttpNotFound(f"Missing parameters: {', '.join(missing_params)}")

        try:
            model = apps.get_model(app_label, module)
            field = next((f for f in model._meta.fields if f.name == field_name), None)
            if not field:
                return render(
                    request,
                    "common/message_fragment.html",
                    {
                        "message": f"Field not found: {field_name}",
                        "variant": "border",
                    },
                )

            unique_file_values = unique_values.get(file_header, [])
            is_choice_field = isinstance(field, CharField) and field.choices
            is_foreign_key = isinstance(field, ForeignKey)

            # Get auto-mappings from session
            auto_choice_mappings = import_data.get("auto_choice_mappings", {})
            auto_fk_mappings = import_data.get("auto_fk_mappings", {})
            context = {
                "field": {
                    "name": field_name,
                    "verbose_name": field.verbose_name.title(),
                    "is_choice_field": is_choice_field,
                    "is_foreign_key": is_foreign_key,
                    "choices": (
                        [
                            {"value": value, "label": label}
                            for value, label in field.choices
                        ]
                        if is_choice_field
                        else []
                    ),
                    "foreign_key_choices": (
                        [
                            {"id": instance.pk, "display": str(instance)}
                            for instance in field.related_model.objects.all()
                        ]
                        if is_foreign_key
                        else []
                    ),
                    "unique_file_values": unique_file_values,
                },
                # add missing mappings for manual selections
                "choice_mappings": import_data.get("choice_mappings", {}).get(
                    field_name, {}
                ),
                "fk_mappings": import_data.get("fk_mappings", {}).get(field_name, {}),
                "auto_choice_mappings": auto_choice_mappings.get(field_name, {}),
                "auto_fk_mappings": auto_fk_mappings.get(field_name, {}),
            }

            return render(request, "import/value_mapping_partial.html", context)
        except Exception as e:
            tb = traceback.format_exc()
            logger.error(tb)
            return render(
                request,
                "common/message_fragment.html",
                {"message": f"Error: {e!s}", "variant": "border"},
            )


@method_decorator(
    permission_required_or_denied("core.can_view_horilla_import"),
    name="dispatch",
)
class UpdateValueMappingStatusView(LoginRequiredMixin, View):
    """View to update value mapping status"""

    def post(self, request, *args, **kwargs):
        """Update value mapping status based on user input"""
        import_data = request.session.get("import_data", {})
        field_name = request.GET.get("field_name")
        slug_value = request.GET.get("slug_value")
        value = request.POST.get(
            f"choice_mapping_{field_name}_{slug_value}"
        ) or request.POST.get(f"fk_mapping_{field_name}_{slug_value}")

        if not all([field_name, slug_value, value]):
            return render(
                request,
                "common/message_fragment.html",
                {"message": _("Error: Missing parameters"), "variant": "badge"},
            )

        try:
            module = import_data.get("module")
            app_label = import_data.get("app_label")
            model = apps.get_model(app_label, module)
            field = next((f for f in model._meta.fields if f.name == field_name), None)

            if not field:
                return render(
                    request,
                    "common/message_fragment.html",
                    {"message": _("Error: Field not found"), "variant": "badge"},
                )

            is_choice_field = isinstance(field, CharField) and field.choices
            is_foreign_key = isinstance(field, ForeignKey)

            if is_choice_field:
                import_data.setdefault("choice_mappings", {})
                import_data["choice_mappings"].setdefault(field_name, {})
                import_data["choice_mappings"][field_name][slug_value] = value
            elif is_foreign_key:
                import_data.setdefault("fk_mappings", {})
                import_data["fk_mappings"].setdefault(field_name, {})
                import_data["fk_mappings"][field_name][slug_value] = str(
                    value
                )  # Store as string
            else:
                return render(
                    request,
                    "common/message_fragment.html",
                    {"message": _("Error: Invalid field type"), "variant": "badge"},
                )

            request.session["import_data"] = import_data
            request.session.modified = True

            return render(
                request,
                "common/message_fragment.html",
                {"message": _("Mapped"), "variant": "success_badge"},
            )
        except Exception as e:
            tb = traceback.format_exc()
            logger.error(tb)
            return render(
                request,
                "common/message_fragment.html",
                {"message": f"Error: {e!s}", "variant": "badge"},
            )


@method_decorator(
    permission_required_or_denied("core.can_view_horilla_import"),
    name="dispatch",
)
class DownloadErrorFileView(LoginRequiredMixin, View):
    """Download error CSV file"""

    def get(self, request, *args, **kwargs):
        """Download the error CSV file"""
        file_path = request.GET.get("file_path")
        if not file_path:
            raise HttpNotFound("File path not provided")

        try:
            if not default_storage.exists(file_path):
                return HttpResponse("File not found", status=404)

            if not file_path.startswith("import_errors/"):
                return HttpResponse("Access denied", status=403)

            with default_storage.open(file_path, "rb") as file:
                response = HttpResponse(file.read(), content_type="text/csv")

                # Extract filename from path
                filename = file_path.split("/")[-1]
                response["Content-Disposition"] = f'attachment; filename="{filename}"'

                return response

        except Exception as e:
            logger.error("Error downloading error file: %s", e)
            return HttpResponse("Error downloading file", status=500)


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied("core.can_view_horilla_import"),
    name="dispatch",
)
class ImportHistoryView(LoginRequiredMixin, HorillaListView):
    """View to display import history records"""

    model = ImportHistory
    view_id = "import-history"
    search_url = reverse_lazy("core:import_history_view")
    main_url = reverse_lazy("core:import_history_view")
    table_width = False
    table_height_as_class = "h-[calc(_100vh_-_330px_)]"

    header_attrs = [
        {"imported_file_path": {"style": "width: 300px;"}},
        {"error_file_path": {"style": "width: 500px;"}},
    ]

    columns = [
        "import_name",
        (_("Module"), "module_verbose_name"),
        "original_filename",
        "status",
        "success_rate",
        "duration_seconds",
        "created_at",
        "created_by",
        (_("Imported File"), "imported_file"),
        (_("Error File"), "error_list"),
    ]


@method_decorator(
    permission_required_or_denied("core.can_view_horilla_import"),
    name="dispatch",
)
class DownloadImportedFileView(LoginRequiredMixin, View):
    """Download the original imported file"""

    def get(self, request, *args, **kwargs):
        """Download the original imported file"""
        file_path = request.GET.get("file_path")

        if not file_path:
            raise HttpNotFound("File path not provided")

        try:
            if not default_storage.exists(file_path):
                return HttpResponse("Imported file not found", status=404)

            # Optional: Add a security check to ensure the file is in a specific directory
            if not file_path.startswith(
                "imports/"
            ):  # Adjust the directory as per your storage structure
                return HttpResponse("Access denied", status=403)

            with default_storage.open(file_path, "rb") as file:
                content = file.read()

                # Determine content type based on file extension
                content_type = (
                    "text/csv"
                    if file_path.endswith(".csv")
                    else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                response = HttpResponse(content, content_type=content_type)
                filename = file_path.split("/")[-1]
                response["Content-Disposition"] = f'attachment; filename="{filename}"'
                return response

        except Exception as e:
            logger.error("Error downloading imported file: %s", e)
            tb = traceback.format_exc()
            logger.error(tb)
            return render(
                request,
                "common/message_fragment.html",
                {"message": f"Error downloading imported file: {e!s}", "variant": "sm"},
                status=500,
            )


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied("core.can_view_horilla_import"),
    name="dispatch",
)
class DownloadTemplateModalView(LoginRequiredMixin, TemplateView):
    """View to show field selection modal for template download"""

    template_name = "import/download_template_modal.html"

    def is_module_allowed(self, module_name, app_label):
        """Check if the module is in the allowed import models list"""
        try:
            import_models = FEATURE_REGISTRY.get("import_models", [])
            for model in import_models:
                if model.__name__ == module_name and model._meta.app_label == app_label:
                    return True
            return False
        except Exception as e:
            logger.error("Error checking if module is allowed: %s", e)
            return False

    def get_context_data(self, **kwargs):
        """Add module, app_label, and import config to context; set error if invalid module."""
        context = super().get_context_data(**kwargs)
        module = self.request.GET.get("module")

        if not module:
            context["error"] = _("Please select a module first")
            return context

        try:
            app_label = self.get_app_label_for_model(module)
            if not app_label:
                context["error"] = _("Invalid module selected")
                return context

            # Validate that the module is in the allowed choices
            if not self.is_module_allowed(module, app_label):
                context["error"] = _("Selected module is not available for import")
                return context

            model_fields = self.get_model_fields(module, app_label)
            context["fields"] = model_fields
            context["module"] = module
            context["app_label"] = app_label
            # Initially all non-mandatory fields are selected
            non_mandatory_fields = [f for f in model_fields if not f["required"]]
            context["select_all"] = len(non_mandatory_fields) > 0
            context["selected_fields"] = {
                f["name"]
                for f in model_fields
                if f["required"] or context["select_all"]
            }
        except Exception as e:
            logger.error("Error getting model fields for template: %s", e)
            context["error"] = _("Error loading fields")

        return context

    def get_app_label_for_model(self, model_name):
        """Get app label for a model name"""
        for app_config in apps.get_app_configs():
            try:
                # Only check installed apps
                if not app_config.models_module:
                    continue
                model = apps.get_model(app_config.label, model_name)
                if model:
                    return app_config.label
            except LookupError:
                continue
        return None

    def get_model_fields(self, module_name, app_label):
        """Get fields from the selected model"""
        try:
            model = apps.get_model(app_label, module_name)
            fields = []
            for field in model._meta.fields:
                if field.name in IMPORT_EXCLUDED_FIELDS:
                    continue
                if not getattr(field, "editable", True):
                    continue

                field_info = {
                    "name": field.name,
                    "verbose_name": field.verbose_name.title(),
                    "required": not field.null and not field.blank,
                }
                fields.append(field_info)
            return fields
        except Exception as e:
            logger.error(
                "Error in get_model_fields (app_label: %s, module: %s): %s",
                app_label,
                module_name,
                e,
            )
            return []


@method_decorator(
    permission_required_or_denied("core.can_view_horilla_import"),
    name="dispatch",
)
class DownloadTemplateView(LoginRequiredMixin, View):
    """View to generate and download template file with selected fields"""

    def is_module_allowed(self, module_name, app_label):
        """Check if the module is in the allowed import models list"""
        try:
            import_models = FEATURE_REGISTRY.get("import_models", [])
            for model in import_models:
                if model.__name__ == module_name and model._meta.app_label == app_label:
                    return True
            return False
        except Exception as e:
            logger.error("Error checking if module is allowed: %s", e)
            return False

    def post(self, request, *args, **kwargs):
        """Generate template file with selected fields"""
        module = request.POST.get("module")
        app_label = request.POST.get("app_label")
        selected_fields = request.POST.getlist("selected_fields")
        file_format = request.POST.get("file_format", "xlsx")

        if not module or not app_label or not selected_fields:
            response = HttpResponse("", status=400)
            response["HX-Refresh"] = "true"
            return response

        try:
            # Validate that the module is in the allowed choices
            if not self.is_module_allowed(module, app_label):
                messages.error(
                    request, _("Selected module is not available for import")
                )
                response = HttpResponse("", status=400)
                response["HX-Refresh"] = "true"
                return response

            # Check if app is installed before trying to get the model
            try:
                apps.get_app_config(app_label)
            except Exception:
                messages.error(
                    request, _("Selected module is not available for import")
                )
                response = HttpResponse("", status=400)
                response["HX-Refresh"] = "true"
                return response

            try:
                model = apps.get_model(app_label, module)
            except Exception:
                messages.error(
                    request, _("Selected module is not available for import")
                )
                response = HttpResponse("", status=400)
                response["HX-Refresh"] = "true"
                return response

            if not model:
                response = HttpResponse("", status=400)
                response["HX-Refresh"] = "true"
                return response

            # Get field verbose names for headers
            field_headers = []
            for field_name in selected_fields:
                try:
                    field = model._meta.get_field(field_name)
                    field_headers.append(field.verbose_name.title())
                except Exception:
                    field_headers.append(field_name)

            # Get model verbose name for filename (lowercase)
            model_verbose_name = model._meta.verbose_name.lower().replace(" ", "_")

            # Generate file
            if file_format == "csv":
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    f'attachment; filename="{model_verbose_name}_template.csv"'
                )
                writer = csv.writer(response)
                writer.writerow(field_headers)
            else:
                # Excel format with styling (like bulk export)
                wb = Workbook()
                ws = wb.active
                ws.title = "Template"

                # Add headers
                ws.append([str(header) for header in field_headers])

                # Style headers (same as bulk export)
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal="center")
                header_fill = PatternFill(
                    start_color="eafb5b", end_color="eafb5b", fill_type="solid"
                )

                for cell in ws[1]:  # First row (headers)
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.fill = header_fill

                # Set column widths (same as bulk export)
                for col in ws.columns:
                    column_letter = col[0].column_letter
                    ws.column_dimensions[column_letter].width = 25

                # Set row height for header
                ws.row_dimensions[1].height = 15

                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                response = HttpResponse(
                    buffer.getvalue(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = (
                    f'attachment; filename="{model_verbose_name}_template.xlsx"'
                )

            return response

        except Exception as e:
            logger.error("Error generating template: %s", e)
            tb = traceback.format_exc()
            logger.error(tb)
            response = HttpResponse("", status=500)
            response["HX-Refresh"] = "true"
            return response

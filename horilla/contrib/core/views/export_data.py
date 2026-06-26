"""
This view handles the methods for export view
"""

import csv

# Standard library imports
import io
import logging
import zipfile
from datetime import date, datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models.fields.related import ManyToManyField

# Third-party imports (Django)
from django.utils.functional import cached_property
from django.utils.text import slugify
from django.views import View
from django.views.generic import TemplateView

# Third-party imports
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from horilla.apps import apps
from horilla.contrib.generics.views import (
    HorillaListView,
    HorillaModalDetailView,
    HorillaSingleDeleteView,
)
from horilla.db.models import ForeignKey
from horilla.registry.feature import FEATURE_REGISTRY
from horilla.shortcuts import render

# First party imports (Horilla)
from horilla.utils import timezone
from horilla.utils.decorators import (
    htmx_required,
    method_decorator,
    permission_required_or_denied,
)
from horilla.utils.translation import gettext_lazy as _
from horilla.web import HttpResponse

# Local imports
from ..models import ExportSchedule
from ..utils import sanitize_export_value

logger = logging.getLogger(__name__)


class ExportView(LoginRequiredMixin, TemplateView):
    """
    Template view for export page
    """

    template_name = "export/export_view.html"

    def get_context_data(self, **kwargs):
        """
        Provide context for export view, including modules,
        selected modules, export format, and scheduled exports.
        """
        context = super().get_context_data(**kwargs)
        context["modules"] = self.get_available_models()

        context["selected_modules"] = self.request.GET.getlist(
            "module"
        ) or self.request.POST.getlist("module")
        context["export_format"] = self.request.GET.get(
            "export_format"
        ) or self.request.POST.get("export_format", "xlsx")
        context["scheduled_exports"] = ExportSchedule.objects.filter(
            user=self.request.user
        ).order_by("-created_at")
        return context

    def get_available_models(self):
        """Return registered export models the user has view permission for."""
        models = []
        try:
            export_models = FEATURE_REGISTRY.get("export_models", [])
            for model in export_models:
                view_perm = f"{model._meta.app_label}.view_{model._meta.model_name}"
                if self.request.user.has_perm(view_perm):
                    models.append(
                        {
                            "name": model.__name__,
                            "label": model._meta.verbose_name.title(),
                            "app_label": model._meta.app_label,
                            "module": model.__module__,
                        }
                    )
        except Exception as e:
            logger.error(e)

        return models

    def post(self, request, *args, **kwargs):
        """
        Handle export requests for selected models.

        Supports CSV, XLSX, and PDF formats.
        If multiple models are selected, bundles them into a ZIP archive.
        """
        selected_models = request.POST.getlist("module")
        export_format = request.POST.get("export_format")

        if not selected_models or not export_format:
            messages.error(
                request, _("Please select at least one model and export format.")
            )
            return self.render_to_response(self.get_context_data())

        if len(selected_models) == 1:
            model_name = selected_models[0]
            model = self.get_model_by_name(model_name)

            if not model:
                messages.error(request, _("Selected module not found."))
                return self.render_to_response(self.get_context_data())

            view_perm = f"{model._meta.app_label}.view_{model._meta.model_name}"
            if not request.user.has_perm(view_perm):
                messages.error(
                    request, _("You do not have permission to export this module.")
                )
                return self.render_to_response(self.get_context_data())

            filename, data = self.export_model_data(model, export_format)

            if not filename or not data:
                messages.error(request, _("Export failed. Please try again."))
                return self.render_to_response(self.get_context_data())

            messages.success(request, _("Export completed successfully."))

            response = HttpResponse(
                data.getvalue(), content_type=self.get_content_type(export_format)
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for model_name in selected_models:
                model = self.get_model_by_name(model_name)
                if not model:
                    continue

                view_perm = f"{model._meta.app_label}.view_{model._meta.model_name}"
                if not request.user.has_perm(view_perm):
                    logger.warning(
                        "Skipping model %s: user %s lacks %s",
                        model_name,
                        request.user.email,
                        view_perm,
                    )
                    continue

                filename, data = self.export_model_data(model, export_format)
                zip_file.writestr(filename, data.getvalue())

        messages.success(request, _("Export completed successfully."))

        zip_buffer.seek(0)
        response = HttpResponse(
            zip_buffer,
            content_type="application/zip",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="export_{export_format}.zip"'
        )
        return response

    def get_content_type(self, export_format):
        """Return the appropriate MIME type for the given export format."""
        content_types = {
            "csv": "text/csv",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "pdf": "application/pdf",
        }
        return content_types.get(export_format, "application/octet-stream")

    def get_model_by_name(self, model_name):
        """Return the Django model class corresponding to the given class name."""
        for model in apps.get_models():
            if model.__name__ == model_name:
                return model
        return None

    def get_export_filename(self, model, export_format):
        """
        Generate a timestamped, safe filename for exporting a Horilla model.
        """

        tz_str = getattr(self.request.user, "time_zone", None)
        user_tz = ZoneInfo(tz_str) if tz_str else ZoneInfo("UTC")
        dt_format = (
            getattr(self.request.user, "date_time_format", None) or "%Y-%m-%d %H:%M:%S"
        )
        now = timezone.now().astimezone(user_tz)

        # filenames cannot safely contain :
        timestamp = now.strftime(dt_format)
        safe_name = slugify(model._meta.verbose_name_plural)
        return f"{safe_name}_export_{timestamp}.{export_format}"

    def _get_export_value(self, obj, field_name, field, user):
        """Delegate to shared helper for use in view and in scheduled export tasks."""
        return get_export_cell_value(obj, field_name, field, user)

    def export_model_data(self, model, export_format):
        """Export all data of a given model in the selected format"""
        queryset = model.objects.all()
        fields = model._meta.fields
        user = getattr(self.request, "user", None)

        field_data = [(str(field.verbose_name), field.name, field) for field in fields]
        column_headers = [fd[0] for fd in field_data]

        data = []
        for obj in queryset:
            row = []
            for _verbose_name, field_name, field in field_data:
                value = self._get_export_value(obj, field_name, field, user)
                row.append(value)
            data.append(row)

        _model_verbose_name = model._meta.verbose_name_plural.lower().replace(" ", "_")
        document_title = f"Exported {model._meta.verbose_name_plural}"

        if export_format == "csv":
            buffer = io.StringIO()
            writer = csv.writer(buffer)
            writer.writerow(column_headers)
            for row in data:
                sanitized_row = [sanitize_export_value(cell) for cell in row]
                writer.writerow(sanitized_row)
            buffer.seek(0)
            filename = self.get_export_filename(model, "csv")
            return f"{filename}", io.BytesIO(buffer.getvalue().encode("utf-8"))

        if export_format == "xlsx":
            wb = Workbook()
            ws = wb.active

            ws.append([str(header) for header in column_headers])

            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal="center")
            header_fill = PatternFill(
                start_color="eafb5b", end_color="eafb5b", fill_type="solid"
            )

            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill

            for col in ws.columns:
                column_letter = col[0].column_letter
                ws.column_dimensions[column_letter].width = 25

            for row in data:
                sanitized_row = [
                    sanitize_export_value(str(cell) if cell is not None else "")
                    for cell in row
                ]
                ws.append(sanitized_row)

            for idx, row in enumerate(ws.rows, 1):
                ws.row_dimensions[idx].height = 15

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            filename = self.get_export_filename(model, "xlsx")
            return f"{filename}", buffer

        if export_format == "pdf":
            from reportlab.lib import colors

            buffer = BytesIO()
            page_size = (letter[1], letter[0])
            width, height = page_size

            c = canvas.Canvas(buffer, pagesize=page_size)
            c.setTitle(document_title)

            title_font_size = 18
            header_font_size = 12
            data_font_size = 10

            start_x = 50
            start_y = height - 100
            min_col_width = 120
            padding = 8
            max_rows_per_page = 7
            max_cols_per_page = 6
            extra_row_spacing = 10

            def wrap_text(text, max_chars):
                text = str(text) if text is not None else ""
                if len(text) <= max_chars:
                    return [text] if text else [""]
                words = text.split()
                lines = []
                current_line = ""
                for word in words:
                    if len(current_line) + len(word) + 1 <= max_chars:
                        current_line += word + " "
                    else:
                        lines.append(current_line.strip())
                        current_line = word + " "
                if current_line:
                    lines.append(current_line.strip())
                return lines if lines else [""]

            column_chunks = [
                column_headers[i : i + max_cols_per_page]
                for i in range(0, len(column_headers), max_cols_per_page)
            ]

            for chunk_idx, chunk_headers in enumerate(column_chunks):
                total_table_width = min(len(chunk_headers) * min_col_width, width - 100)
                col_width = (
                    total_table_width / len(chunk_headers) if chunk_headers else 100
                )
                max_chars_per_line = int(col_width // (header_font_size * 0.5))

                rows_drawn = 0
                row_start = 0

                while row_start < len(data):
                    c.setFont("Helvetica-Bold", title_font_size)
                    column_range = f"Columns {(chunk_idx * max_cols_per_page) + 1} to {min((chunk_idx + 1) * max_cols_per_page, len(column_headers))}"
                    c.drawCentredString(
                        width / 2, height - 50, f"{document_title} ({column_range})"
                    )

                    c.setFont("Helvetica-Bold", header_font_size)
                    c.setFillColor(colors.black)
                    header_y = start_y
                    max_header_lines = 1
                    for i, header in enumerate(chunk_headers):
                        wrapped_header = wrap_text(header, max_chars_per_line)
                        max_header_lines = max(max_header_lines, len(wrapped_header))

                    header_height = max_header_lines * (header_font_size + 2) + 15
                    c.setFillColor(colors.lightgrey)
                    c.rect(
                        start_x,
                        header_y - header_height + 5,
                        total_table_width,
                        header_height,
                        fill=1,
                        stroke=0,
                    )

                    c.setFont("Helvetica-Bold", header_font_size)
                    c.setFillColor(colors.black)
                    for i, header in enumerate(chunk_headers):
                        x = start_x + i * col_width + padding
                        wrapped_header = wrap_text(header, max_chars_per_line)
                        total_text_height = len(wrapped_header) * (header_font_size + 2)
                        y_offset = (header_height - total_text_height) / 2 + 3
                        for line in wrapped_header:
                            c.drawString(x, header_y - y_offset, line)
                            y_offset += header_font_size + 2

                    c.setFont("Helvetica", data_font_size)
                    y = header_y - header_height - 10
                    rows_drawn = 0

                    for row_idx in range(row_start, len(data)):
                        if rows_drawn >= max_rows_per_page:
                            break

                        start_col = chunk_idx * max_cols_per_page
                        end_col = min(
                            (chunk_idx + 1) * max_cols_per_page, len(data[row_idx])
                        )
                        row = data[row_idx][start_col:end_col]

                        max_lines_in_row = 1
                        for value in row:
                            wrapped_value = wrap_text(value, max_chars_per_line)
                            max_lines_in_row = max(max_lines_in_row, len(wrapped_value))
                        row_height = (
                            max_lines_in_row * (data_font_size + 2) + extra_row_spacing
                        )
                        total_text_height = max_lines_in_row * (data_font_size + 2)
                        text_y_offset = (row_height - total_text_height) / 2 + 9

                        if rows_drawn % 2 == 0:
                            c.setFillColor(colors.whitesmoke)
                            c.rect(
                                start_x,
                                y - row_height,
                                total_table_width,
                                row_height,
                                fill=1,
                                stroke=0,
                            )

                        for i, value in enumerate(row):
                            wrapped_value = wrap_text(value, max_chars_per_line)
                            x = start_x + i * col_width + padding
                            y_offset = text_y_offset
                            for line in wrapped_value:
                                c.setFillColor(colors.black)
                                c.drawString(x, y - y_offset, line)
                                y_offset += data_font_size + 2

                        y -= row_height
                        rows_drawn += 1

                    row_start += max_rows_per_page
                    c.showPage()

            c.save()
            buffer.seek(0)
            filename = self.get_export_filename(model, "pdf")
            return f"{filename}", buffer

        return None, None


def get_export_cell_value(obj, field_name, field, user):
    """
    Get cell value for export: display labels for choices/CountryField,
    user-formatted date/datetime, and proper handling for FK/M2M.
    Shared by ExportView and scheduled export tasks.
    """
    try:
        # Prefer display value for choice fields and CountryField
        display_method_name = f"get_{field_name}_display"
        if hasattr(obj, display_method_name):
            display_method = getattr(obj, display_method_name)
            if callable(display_method):
                value = display_method()
            else:
                value = getattr(obj, field_name, "")
        else:
            value = getattr(obj, field_name, "")
        if isinstance(field, ForeignKey):
            value = str(getattr(value, "username", value)) if value else ""
        elif isinstance(field, ManyToManyField):
            value = ", ".join(str(item) for item in value.all()) if value else ""
        # Format date/datetime with user's chosen format and timezone
        if user and value is not None:
            if isinstance(value, datetime):
                if getattr(user, "time_zone", None):
                    try:
                        user_tz = ZoneInfo(user.time_zone)
                        if timezone.is_naive(value):
                            value = timezone.make_aware(
                                value, timezone.get_default_timezone()
                            )
                        value = value.astimezone(user_tz)
                    except Exception:
                        pass
                fmt = getattr(user, "date_time_format", None) or "%Y-%m-%d %H:%M:%S"
                try:
                    value = value.strftime(fmt)
                except Exception:
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(value, date) and not isinstance(value, datetime):
                fmt = getattr(user, "date_format", None) or "%Y-%m-%d"
                try:
                    value = value.strftime(fmt)
                except Exception:
                    value = value.strftime("%Y-%m-%d")
        value_str = str(value) if value is not None else ""
        return sanitize_export_value(value_str)
    except Exception as e:
        logger.error("Error retrieving field %s: %s", field_name, str(e))
        return ""


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied("core.can_view_horilla_export"),
    name="dispatch",
)
class ExportScheduleModalView(LoginRequiredMixin, View):
    """
    Modal view for displaying a scheduled export form.

    Handles fetching a schedule by ID or initializing a new schedule form
    with selected modules, export format, and frequency.
    """

    def get(self, request):
        """
        ExportScheduleModalView get request method
        """
        schedule_id = request.GET.get("id")
        schedule = None
        if schedule_id:
            try:
                schedule = ExportSchedule.objects.get(pk=schedule_id, user=request.user)
                modules = schedule.modules
                export_format = schedule.export_format
                selected_frequency_option = schedule.frequency
                start_date = (
                    schedule.start_date.strftime("%Y-%m-%d")
                    if schedule.start_date
                    else ""
                )
                end_date = (
                    schedule.end_date.strftime("%Y-%m-%d") if schedule.end_date else ""
                )
                form_data = {
                    "weekday": schedule.weekday,
                    "day_of_month": schedule.day_of_month,
                    "yearly_day_of_month": schedule.yearly_day_of_month,
                    "yearly_month": schedule.yearly_month,
                    "start_date": start_date,
                    "end_date": end_date,
                }
            except ExportSchedule.DoesNotExist:
                messages.error(request, _("Schedule not found."))
                return HttpResponse(
                    "<script>closeModal();$('#reloadScheduleListButton').click();</script>"
                )
        else:
            modules = request.GET.getlist("module")
            if not modules:
                messages.error(
                    request, _("Please select at least one module before scheduling.")
                )
                return HttpResponse(
                    "<script>$('#reloadButton').click();$('#reloadMessagesButton').click();closeModal();</script>"
                )

            export_format = request.GET.get("export_format", "xlsx")
            selected_frequency_option = request.GET.get("frequency", "daily")
            form_data = {}
        context = {
            "selected_modules": modules,
            "schedule": schedule,
            "schedule_id": schedule_id,
            "selected_format": export_format,
            "selected_frequency_option": selected_frequency_option,
            "form_data": form_data,
            "weekdays": [
                "monday",
                "tuesday",
                "wednesday",
                "thursday",
                "friday",
                "saturday",
                "sunday",
            ],
            "days_in_month": list(range(1, 32)),
            "months": [
                (1, _("January")),
                (2, _("February")),
                (3, _("March")),
                (4, _("April")),
                (5, _("May")),
                (6, _("June")),
                (7, _("July")),
                (8, _("August")),
                (9, _("September")),
                (10, _("October")),
                (11, _("November")),
                (12, _("December")),
            ],
        }
        return render(request, "export/schedule_modal.html", context)


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied("core.can_view_horilla_export"),
    name="dispatch",
)
class ExportScheduleCreateView(LoginRequiredMixin, View):
    """
    Handles creation and updating of scheduled exports.

    Validates the submitted schedule form, applies frequency rules,
    and saves or updates the ExportSchedule instance.
    """

    def post(self, request):
        """
        ExportScheduleCreateView post request method
        """
        field_errors = {}
        non_field_error = None

        try:
            schedule_id = request.POST.get("schedule_id")

            modules = request.POST.getlist("module")
            export_format = request.POST["export_format"]
            frequency = request.POST["frequency"]

            day_of_month = None
            weekday = None
            yearly_day_of_month = None
            yearly_month = None

            if frequency == "weekly":
                weekday = request.POST.get("weekday")
            elif frequency == "monthly":
                day_of_month = request.POST.get("day_of_month")
            elif frequency == "yearly":
                yearly_day_of_month = request.POST.get("yearly_day_of_month")
                yearly_month = request.POST.get("yearly_month")

            start_date = request.POST.get("start_date")
            end_date = request.POST.get("end_date") or None

            # Validate start_date
            if not start_date:
                field_errors["start_date"] = _("Start date is required.")
            else:
                try:
                    start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()

                    today = timezone.now().date()
                    if start_date_obj < today:
                        field_errors["start_date"] = _(
                            "Start date cannot be in the past."
                        )
                except ValueError:
                    field_errors["start_date"] = _("Invalid date format.")

            # Validate end_date if provided
            if end_date:
                try:
                    end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
                    if start_date and not field_errors.get("start_date"):
                        if end_date_obj < start_date_obj:
                            field_errors["end_date"] = _(
                                "End date must be after start date."
                            )
                except ValueError:
                    field_errors["end_date"] = _("Invalid date format.")

            # Validate frequency-specific fields
            if frequency == "yearly":
                if not yearly_day_of_month or not yearly_month:
                    non_field_error = _(
                        "Day of month and month are required for yearly schedule."
                    )
                    raise ValueError(non_field_error)

            if frequency == "weekly" and not weekday:
                field_errors["weekday"] = _(
                    "Day of week is required for weekly schedule."
                )

            if frequency == "monthly" and not day_of_month:
                field_errors["day_of_month"] = _(
                    "Day of month is required for monthly schedule."
                )

            if field_errors:
                raise ValueError("Validation failed")

            if schedule_id:
                schedule = ExportSchedule.objects.get(pk=schedule_id, user=request.user)
                schedule.modules = modules
                schedule.export_format = export_format
                schedule.frequency = frequency
                schedule.day_of_month = day_of_month
                schedule.weekday = weekday
                schedule.yearly_day_of_month = yearly_day_of_month
                schedule.yearly_month = yearly_month
                schedule.start_date = start_date
                schedule.end_date = end_date
                schedule.updated_by = request.user
                schedule.save()

                success_message = _("Scheduled export updated successfully.")
            else:
                schedule = ExportSchedule(
                    user=request.user,
                    modules=modules,
                    export_format=export_format,
                    frequency=frequency,
                    day_of_month=day_of_month,
                    weekday=weekday,
                    yearly_day_of_month=yearly_day_of_month,
                    yearly_month=yearly_month,
                    start_date=start_date,
                    end_date=end_date,
                    created_by=request.user,
                    updated_by=request.user,
                    company=getattr(self.request, "active_company", None),
                )
                schedule.save()

                success_message = _("Scheduled export created successfully.")

            messages.success(request, success_message)
            return HttpResponse(
                """
                                <script>
                                    closeModal();
                                    setTimeout(function() {
                                        $('#reloadScheduleListButton').click();
                                        $('#reloadButton').click();
                                        $('#reloadMessagesButton').click();
                                    }, 200);
                                </script>
                                """
            )

        except Exception as e:
            if not field_errors and not non_field_error:
                non_field_error = str(e).strip("[]'")

            context = {
                "selected_modules": request.POST.getlist("module"),
                "selected_format": request.POST.get("export_format", "xlsx"),
                "selected_frequency_option": request.POST.get("frequency", "daily"),
                "field_errors": field_errors,
                "non_field_error": non_field_error,
                "weekdays": [
                    "monday",
                    "tuesday",
                    "wednesday",
                    "thursday",
                    "friday",
                    "saturday",
                    "sunday",
                ],
                "days_in_month": list(range(1, 32)),
                "months": [
                    (1, _("January")),
                    (2, _("February")),
                    (3, _("March")),
                    (4, _("April")),
                    (5, _("May")),
                    (6, _("June")),
                    (7, _("July")),
                    (8, _("August")),
                    (9, _("September")),
                    (10, _("October")),
                    (11, _("November")),
                    (12, _("December")),
                ],
                "form_data": {
                    "weekday": request.POST.get("weekday"),
                    "day_of_month": request.POST.get("day_of_month"),
                    "yearly_day_of_month": request.POST.get("yearly_day_of_month"),
                    "yearly_month": request.POST.get("yearly_month"),
                    "start_date": request.POST.get("start_date"),
                    "end_date": request.POST.get("end_date"),
                },
            }
            return render(request, "export/schedule_modal.html", context)


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied("core.can_view_horilla_export"),
    name="dispatch",
)
class ScheduleExportListView(LoginRequiredMixin, HorillaListView):
    """
    List view for displaying all scheduled exports for the logged-in user.

    Columns include modules, export format, frequency, schedule details,
    and last executed timestamp.
    """

    model = ExportSchedule
    view_id = "Schedule_Export_List"
    table_width = False
    bulk_select_option = False
    list_column_visibility = False
    enable_sorting = False
    no_record_section = False
    store_ordered_ids = True

    columns = [
        (_("Modules"), "module_names_display"),
        "export_format",
        "frequency",
        (_("Schedule Details"), "frequency_display"),
        (_("Last Executed On"), "last_executed"),
    ]

    @cached_property
    def col_attrs(self):
        """
        Add HTMX attributes to the "Modules" column for opening the detail modal on click.
        """
        query_string = self.request.session.get(self.ordered_ids_key, [])
        attrs = {
            "hx-get": f"{{get_detail_url}}?instance_ids={query_string}",
            "hx-target": "#detailModalBox",
            "hx-swap": "innerHTML",
            "hx-push-url": "false",
            "hx-on:click": "openDetailModal();",
            "style": "cursor:pointer",
            "class": "hover:text-primary-600",
        }
        return [{"module_names_display": {**attrs}}]

    actions = [
        {
            "action": "Edit",
            "src": "assets/icons/edit.svg",
            "img_class": "w-4 h-4",
            "attrs": """
                    hx-get="{get_edit_url}?id={id}"
                    hx-target="#modalBox"
                    hx-swap="innerHTML"
                    onclick="openModal()"
                    """,
        },
        {
            "action": "Delete",
            "src": "assets/icons/a4.svg",
            "img_class": "w-4 h-4",
            "attrs": """
                    hx-post="{get_delete_url}"
                    hx-target="#deleteModeBox"
                    hx-swap="innerHTML"
                    hx-trigger="click"
                    hx-vals='{{"check_dependencies": "true"}}'
                    onclick="openDeleteModeModal()"
                """,
        },
    ]


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied("core.can_view_horilla_export", modal=True),
    name="dispatch",
)
class ScheduleExportDeleteView(LoginRequiredMixin, HorillaSingleDeleteView):
    """
    Handles deletion of a scheduled export.

    Triggers frontend refresh actions to reload the schedule list and messages.
    """

    model = ExportSchedule

    def get_post_delete_response(self):
        return HttpResponse(
            "<script>$('#reloadScheduleListButton').click();$('#reloadButton').click();$('#reloadMessagesButton').click();closeModal();closeDetailModal();</script>"
        )


@method_decorator(htmx_required, name="dispatch")
@method_decorator(
    permission_required_or_denied("core.can_view_horilla_export"),
    name="dispatch",
)
class ScheduleExportDetailView(LoginRequiredMixin, HorillaModalDetailView):
    """
    Detail view for a scheduled export.
    """

    model = ExportSchedule
    title = _("Scheduled Export Details")

    header = {
        "title": "module_names_display",
        "subtitle": "frequency_display",
        "avatar": "",
    }

    body = []

    def get_queryset(self):
        queryset = ExportSchedule.all_objects.filter(user=self.request.user)
        instance_ids = self.request.session.get(self.ordered_ids_key, [])
        if instance_ids:
            queryset = queryset.filter(pk__in=instance_ids)
        return queryset

    def get_body_fields(self):
        from horilla.core.exceptions import FieldDoesNotExist

        instance = self.instance or self.object
        fields = [(_("Modules"), "module_names_display"), "export_format", "frequency"]

        if instance.frequency == "weekly":
            fields.append("weekday")
        elif instance.frequency == "monthly":
            fields.append("day_of_month")
        elif instance.frequency == "yearly":
            fields += ["yearly_month", "yearly_day_of_month"]

        fields += ["start_date", "end_date", "last_run"]

        normalized = []
        blank = self.model()
        for entry in fields:
            if isinstance(entry, tuple):
                normalized.append(entry)
                continue
            try:
                label = blank._meta.get_field(entry).verbose_name
            except FieldDoesNotExist:
                label = entry.replace("_", " ").title()
            normalized.append((label, entry))
        return normalized

    actions = [
        {
            "action": "Edit",
            "src": "assets/icons/edit_white.svg",
            "img_class": "w-3 h-3 flex gap-4 filter brightness-0 invert",
            "permission": "core.can_view_horilla_export",
            "attrs": """
                class="w-24 justify-center px-4 py-2 bg-primary-600 text-white rounded-md text-xs flex items-center gap-2 hover:bg-primary-800 transition duration-300 disabled:cursor-not-allowed"
                hx-get="{get_edit_url}?id={id}"
                hx-target="#modalBox"
                hx-swap="innerHTML"
                onclick="openModal();"
            """,
        },
        {
            "action": "Delete",
            "src": "assets/icons/a4.svg",
            "img_class": "svg-themed w-3 h-3",
            "permission": "core.can_view_horilla_export",
            "attrs": """
                class="w-24 justify-center px-4 py-2 bg-[white] rounded-md text-xs flex items-center gap-2 border border-primary-500 hover:border-primary-600 transition duration-300 disabled:cursor-not-allowed text-primary-600"
                hx-post="{get_delete_url}"
                hx-target="#deleteModeBox"
                hx-swap="innerHTML"
                hx-trigger="click"
                hx-vals='{{"check_dependencies": "true"}}'
                onclick="openDeleteModeModal()"
            """,
        },
    ]

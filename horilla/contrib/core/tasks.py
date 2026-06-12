"""Celery tasks for scheduled data exports and cleanup operations.

This module provides background tasks for:
- Processing scheduled exports based on frequency (daily, weekly, monthly, yearly)
- Generating export files in multiple formats (CSV, XLSX, PDF)
- Sending export files via email
- Cleaning up expired export schedules
"""

# Standard library imports
import csv
import io
import logging
import zipfile
from datetime import timedelta
from io import BytesIO

# Third-party imports
from celery import shared_task

# Third-party imports (Django)
from django.core.mail import get_connection
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from horilla.apps import apps
from horilla.contrib.utils.middlewares import _thread_local
from horilla.db import models

# First party imports (Horilla)
from horilla.utils import timezone
from horilla.utils.translation import gettext as _

from .utils import sanitize_export_value

# Local imports
from .views.export_data import get_export_cell_value

logger = logging.getLogger(__name__)


@shared_task
def process_scheduled_exports():
    """
    Main task to process all due scheduled exports.
    This should be run periodically (e.g., every hour or every 15 minutes).
    """
    from .models import ExportSchedule

    now = timezone.now()
    current_date = now.date()
    current_weekday = now.strftime("%A").lower()

    logger.info("=== process_scheduled_exports running at %s ===", now)
    logger.info("Current: date=%s, weekday=%s", current_date, current_weekday)

    # Get all active schedules
    schedules = ExportSchedule.objects.filter(start_date__lte=current_date).filter(
        models.Q(end_date__isnull=True) | models.Q(end_date__gte=current_date)
    )

    logger.info("Found %s active schedules", schedules.count())

    for schedule in schedules:
        try:
            logger.info(
                "Checking schedule %s: frequency=%s, last_run=%s",
                schedule.id,
                schedule.frequency,
                schedule.last_run,
            )
            if should_run_schedule(schedule, current_date, current_weekday):
                logger.info(
                    "✓ Schedule %s should run - triggering execute_scheduled_export",
                    schedule.id,
                )
                execute_scheduled_export.delay(schedule.id)
            else:
                logger.info("✗ Schedule %s should NOT run yet", schedule.id)
        except Exception as e:
            logger.error("Error checking schedule %s: %s", schedule.id, e)
            logger.exception(e)

    return f"Processed {schedules.count()} schedules"


def should_run_schedule(schedule, current_date, current_weekday):
    """
    Determine if a schedule should run based on frequency and last_run date.
    Ensures the schedule only runs once per frequency period.
    """
    last_run = schedule.last_run

    if last_run is None:
        logger.info("Schedule has never run before")
        return check_frequency_match(schedule, current_date, current_weekday)

    days_since_last_run = (current_date - last_run).days
    logger.info("Days since last run: %s", days_since_last_run)

    if schedule.frequency == "daily":
        should_run = days_since_last_run >= 1
        logger.info(
            "  Daily schedule: days_since_last_run=%s, should_run=%s",
            days_since_last_run,
            should_run,
        )
        return should_run and check_frequency_match(
            schedule, current_date, current_weekday
        )

    if schedule.frequency == "weekly":
        should_run = days_since_last_run >= 7 and current_weekday == schedule.weekday
        logger.info(
            "  Weekly schedule: days_since_last_run=%s, weekday_match=%s, should_run=%s",
            days_since_last_run,
            current_weekday == schedule.weekday,
            should_run,
        )
        return should_run

    if schedule.frequency == "monthly":
        month_changed = current_date.year > last_run.year or (
            current_date.year == last_run.year and current_date.month > last_run.month
        )
        day_matches = current_date.day == schedule.day_of_month
        should_run = month_changed and day_matches
        logger.info(
            "  Monthly schedule: month_changed=%s, day_matches=%s, should_run=%s",
            month_changed,
            day_matches,
            should_run,
        )
        return should_run

    if schedule.frequency == "yearly":
        year_changed = current_date.year > last_run.year
        date_matches = (
            current_date.day == schedule.yearly_day_of_month
            and current_date.month == schedule.yearly_month
        )
        should_run = year_changed and date_matches
        logger.info(
            "  Yearly schedule: year_changed=%s, date_matches=%s, should_run=%s",
            year_changed,
            date_matches,
            should_run,
        )
        return should_run

    logger.info("  ✗ Unknown frequency: %s", schedule.frequency)
    return False


def check_frequency_match(schedule, current_date, current_weekday):
    """
    Check if the current date matches the schedule's frequency requirements.
    Used for first-time runs.
    """
    if schedule.frequency == "daily":
        logger.info("  ✓ Daily schedule - will run")
        return True

    if schedule.frequency == "weekly":
        match = current_weekday == schedule.weekday
        logger.info(
            "  Weekly: current=%s, required=%s, match=%s",
            current_weekday,
            schedule.weekday,
            match,
        )
        return match

    if schedule.frequency == "monthly":
        match = current_date.day == schedule.day_of_month
        logger.info(
            "  Monthly: current_day=%s, required=%s, match=%s",
            current_date.day,
            schedule.day_of_month,
            match,
        )
        return match

    if schedule.frequency == "yearly":
        match = (
            current_date.day == schedule.yearly_day_of_month
            and current_date.month == schedule.yearly_month
        )
        logger.info(
            "  Yearly: current=%s/%s, required=%s/%s, match=%s",
            current_date.month,
            current_date.day,
            schedule.yearly_month,
            schedule.yearly_day_of_month,
            match,
        )
        return match

    return False


@shared_task
def execute_scheduled_export(schedule_id):
    """
    Execute a specific scheduled export and send via email.
    Updates the last_run timestamp after successful execution.
    """
    from .models import ExportSchedule

    logger.info(
        "=== Starting execute_scheduled_export for schedule_id: %s ===",
        schedule_id,
    )

    try:
        schedule = ExportSchedule.objects.get(id=schedule_id)
        logger.info(
            "Found schedule: user=%s, company=%s, modules=%s",
            schedule.user.email,
            schedule.company,
            schedule.modules,
        )
    except ExportSchedule.DoesNotExist:
        logger.error("ExportSchedule %s not found", schedule_id)
        return

    if not schedule.user.has_perm("core.can_view_horilla_export"):
        logger.warning(
            "Skipping schedule %s: user %s no longer has export permission",
            schedule_id,
            schedule.user.email,
        )
        return

    try:
        logger.info("Generating export files for %s modules", len(schedule.modules))
        export_files = generate_export_files(
            schedule.modules, schedule.export_format, user=schedule.user
        )

        if not export_files:
            logger.error("No files generated for schedule %s", schedule_id)
            return

        logger.info("Generated %s export files", len(export_files))

        logger.info("Sending email to %s", schedule.user.email)
        send_export_email(
            user=schedule.user,
            export_format=schedule.export_format,
            export_files=export_files,
            modules=schedule.modules,
            company=schedule.company,
        )

        schedule.last_run = timezone.now().date()
        schedule.save(update_fields=["last_run"])
        logger.info("Updated last_run to %s", schedule.last_run)

        logger.info("=== Successfully executed schedule %s ===", schedule_id)

    except Exception as e:
        logger.error("=== Error executing schedule %s: %s ===", schedule_id, e)
        logger.exception(e)


def generate_export_files(module_names, export_format, user=None):
    """
    Generate export files for the specified modules.
    Returns a list of tuples: (filename, file_data).
    user: optional; when set, choice/date/datetime fields use display values and user format/timezone.
    """
    export_files = []

    for model_name in module_names:
        try:
            model = get_model_by_name(model_name)
            if not model:
                logger.warning("Model %s not found", model_name)
                continue

            if user is not None:
                view_perm = f"{model._meta.app_label}.view_{model._meta.model_name}"
                if not user.has_perm(view_perm):
                    logger.warning(
                        "Skipping model %s: user %s lacks %s",
                        model_name,
                        user.email,
                        view_perm,
                    )
                    continue

            filename, data = export_model_data(model, export_format, user=user)
            if filename and data:
                export_files.append((filename, data))
        except Exception as e:
            logger.error("Error exporting model %s: %s", model_name, e)
            continue

    return export_files


def get_model_by_name(model_name):
    """Find model by class name."""
    for model in apps.get_models():
        if model.__name__ == model_name:
            return model
    return None


def export_model_data(model, export_format, user=None):
    """
    Export all data of a given model in the selected format.
    Returns tuple of (filename, BytesIO buffer).
    user: optional; when set, choice/date/datetime use display values and user format/timezone.
    """
    queryset = model.objects.all()
    fields = model._meta.fields

    field_data = [(str(field.verbose_name), field.name, field) for field in fields]
    column_headers = [fd[0] for fd in field_data]

    data = []
    for obj in queryset:
        row = []
        for _verbose_name, field_name, field in field_data:
            value = get_export_cell_value(obj, field_name, field, user)
            row.append(value)
        data.append(row)

    if export_format == "csv":
        return export_to_csv(model, column_headers, data)
    if export_format == "xlsx":
        return export_to_xlsx(model, column_headers, data)
    if export_format == "pdf":
        return export_to_pdf(model, column_headers, data)

    return None, None


def export_to_csv(model, headers, data):
    """Export data to CSV format."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in data:
        sanitized_row = [sanitize_export_value(cell) for cell in row]
        writer.writerow(sanitized_row)
    buffer.seek(0)
    return f"{model.__name__}_export.csv", io.BytesIO(buffer.getvalue().encode("utf-8"))


def export_to_xlsx(model, headers, data):
    """Export data to Excel format."""
    wb = Workbook()
    ws = wb.active

    ws.append([str(header) for header in headers])

    # Style headers
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")
    header_fill = PatternFill(
        start_color="eafb5b", end_color="eafb5b", fill_type="solid"
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill

    # Set column widths
    for col in ws.columns:
        column_letter = col[0].column_letter
        ws.column_dimensions[column_letter].width = 25

    # Add data
    for row in data:
        sanitized_row = [
            sanitize_export_value(str(cell) if cell is not None else "") for cell in row
        ]
        ws.append(sanitized_row)

    # Set row heights
    for idx, row in enumerate(ws.rows, 1):
        ws.row_dimensions[idx].height = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return f"{model.__name__}_export.xlsx", buffer


def export_to_pdf(model, headers, data):
    """Export data to PDF format."""
    buffer = BytesIO()
    page_size = (letter[1], letter[0])
    width, height = page_size

    c = canvas.Canvas(buffer, pagesize=page_size)
    document_title = f"Exported {model._meta.verbose_name_plural}"
    c.setTitle(document_title)

    # PDF generation parameters
    title_font_size = 18
    header_font_size = 12
    data_font_size = 10
    start_x = 50
    start_y = height - 100
    min_col_width = 120
    padding = 8
    max_rows_per_page = 7
    max_cols_per_page = 6
    extra_row_spacing = 10

    def wrap_text(text, max_chars):
        text = str(text) if text is not None else ""
        if len(text) <= max_chars:
            return [text] if text else [""]
        words = text.split()
        lines = []
        current_line = ""
        for word in words:
            if len(current_line) + len(word) + 1 <= max_chars:
                current_line += word + " "
            else:
                lines.append(current_line.strip())
                current_line = word + " "
        if current_line:
            lines.append(current_line.strip())
        return lines if lines else [""]

    column_chunks = [
        headers[i : i + max_cols_per_page]
        for i in range(0, len(headers), max_cols_per_page)
    ]

    for chunk_idx, chunk_headers in enumerate(column_chunks):
        total_table_width = min(len(chunk_headers) * min_col_width, width - 100)
        col_width = total_table_width / len(chunk_headers) if chunk_headers else 100
        max_chars_per_line = int(col_width // (header_font_size * 0.5))

        row_start = 0

        while row_start < len(data):
            # Draw title
            c.setFont("Helvetica-Bold", title_font_size)
            column_range = f"Columns {(chunk_idx * max_cols_per_page) + 1} to {min((chunk_idx + 1) * max_cols_per_page, len(headers))}"
            c.drawCentredString(
                width / 2, height - 50, f"{document_title} ({column_range})"
            )

            # Draw headers
            c.setFont("Helvetica-Bold", header_font_size)
            header_y = start_y
            max_header_lines = max(
                [len(wrap_text(h, max_chars_per_line)) for h in chunk_headers]
            )
            header_height = max_header_lines * (header_font_size + 2) + 15

            c.setFillColor(colors.lightgrey)
            c.rect(
                start_x,
                header_y - header_height + 5,
                total_table_width,
                header_height,
                fill=1,
                stroke=0,
            )

            c.setFillColor(colors.black)
            for i, header in enumerate(chunk_headers):
                x = start_x + i * col_width + padding
                wrapped_header = wrap_text(header, max_chars_per_line)
                y_offset = (
                    header_height - len(wrapped_header) * (header_font_size + 2)
                ) / 2 + 3
                for line in wrapped_header:
                    c.drawString(x, header_y - y_offset, line)
                    y_offset += header_font_size + 2

            # Draw data rows
            c.setFont("Helvetica", data_font_size)
            y = header_y - header_height - 10
            rows_drawn = 0

            for row_idx in range(row_start, len(data)):
                if rows_drawn >= max_rows_per_page:
                    break

                start_col = chunk_idx * max_cols_per_page
                end_col = min((chunk_idx + 1) * max_cols_per_page, len(data[row_idx]))
                row = data[row_idx][start_col:end_col]

                max_lines_in_row = max(
                    [len(wrap_text(v, max_chars_per_line)) for v in row]
                )
                row_height = max_lines_in_row * (data_font_size + 2) + extra_row_spacing

                if rows_drawn % 2 == 0:
                    c.setFillColor(colors.whitesmoke)
                    c.rect(
                        start_x,
                        y - row_height,
                        total_table_width,
                        row_height,
                        fill=1,
                        stroke=0,
                    )

                for i, value in enumerate(row):
                    wrapped_value = wrap_text(value, max_chars_per_line)
                    x = start_x + i * col_width + padding
                    text_y_offset = (
                        row_height - len(wrapped_value) * (data_font_size + 2)
                    ) / 2 + 9
                    for line in wrapped_value:
                        c.setFillColor(colors.black)
                        c.drawString(x, y - text_y_offset, line)
                        text_y_offset += data_font_size + 2

                y -= row_height
                rows_drawn += 1

            row_start += max_rows_per_page
            c.showPage()

    c.save()
    buffer.seek(0)
    return f"{model.__name__}_export.pdf", buffer


def send_export_email(user, export_format, export_files, modules, company=None):
    """
    Send email with export files attached using HorillaDefaultMailBackend.
    """
    logger.info("Starting email send for user: %s, company: %s", user.email, company)

    subject = _("Scheduled Export - {}").format(
        timezone.now().strftime("%Y-%m-%d %H:%M")
    )

    module_names = ", ".join(modules)

    body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
    </head>
    <body style="font-family: Arial, sans-serif; background-color: #f4f4f4; margin: 0; padding: 20px;">
        <div style="max-width: 650px; margin: auto; background: white; border-radius: 12px; padding: 35px; box-shadow: 0 4px 12px rgba(0,0,0,0.08);">

            <!-- Header -->
            <h2 style="color: #000000; text-align: center; font-size: 24px; margin-bottom: 25px;">
                Scheduled Export Completed
            </h2>

            <!-- Intro -->
            <p style="font-size: 14px; color: #333; line-height: 1.6;">
                Hi {user},
            </p>

            <p style="font-size: 14px; color: #333; line-height: 1.6;">
                Your scheduled export has been completed successfully. The exported data is now ready and attached to this email.
            </p>

            <!-- Info Box -->
            <div style="margin: 20px 0; padding: 15px; background: #fdf2f1; border-left: 4px solid #e54f38; border-radius: 6px;">
                <p style="margin: 6px 0; font-size: 14px; color: #333;">
                    📊 <strong>Export Details:</strong>
                </p>
                <p style="margin: 6px 0; font-size: 14px; color: #333;">
                    • <strong>Modules:</strong> {module_names}
                </p>
                <p style="margin: 6px 0; font-size: 14px; color: #333;">
                    • <strong>Format:</strong> {export_format.upper()}
                </p>
                <p style="margin: 6px 0; font-size: 14px; color: #333;">
                    • <strong>Generated:</strong> {timezone.now().strftime("%B %d, %Y at %I:%M %p")}
                </p>
            </div>

            <p style="font-size: 14px; color: #333; line-height: 1.6;">
                Please find the exported file(s) attached to this email. You can download and use them as needed.
            </p>

            <!-- Footer -->
            <hr style="margin: 30px 0; border: none; border-top: 1px solid #eee;">

            <p style="font-size: 12px; color: #888; text-align: center; line-height: 1.5;">
                This is an automated message from <strong>{company}</strong>.<br>
                If you did not schedule this export, please contact your administrator.
            </p>
        </div>
    </body>
    </html>
    """

    try:
        from django.core.mail import EmailMultiAlternatives

        connection = get_connection(
            backend="horilla.contrib.mail.backends.HorillaDefaultMailBackend",
            fail_silently=False,
        )

        email = EmailMultiAlternatives(
            subject=subject, body="", to=[user.email], connection=connection
        )
        email.attach_alternative(body, "text/html")

        # Attach files
        if len(export_files) == 1:
            filename, file_data = export_files[0]
            email.attach(
                filename, file_data.getvalue(), get_content_type(export_format)
            )
        else:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                for filename, file_data in export_files:
                    zip_file.writestr(filename, file_data.getvalue())

            zip_buffer.seek(0)
            zip_filename = (
                f"export_{export_format}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.zip"
            )
            email.attach(zip_filename, zip_buffer.getvalue(), "application/zip")

        email.send(fail_silently=False)
        logger.info("Export email sent successfully to %s", user.email)

    except Exception as e:
        logger.error("Failed to send export email to %s: %s", user.email, e)
        logger.exception(e)
        raise
    finally:
        # Clean up thread local
        if hasattr(_thread_local, "request"):
            delattr(_thread_local, "request")


def get_content_type(export_format):
    """Get content type based on export format."""
    content_types = {
        "csv": "text/csv",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "pdf": "application/pdf",
    }
    return content_types.get(export_format, "application/octet-stream")


@shared_task
def cleanup_old_schedules():
    """
    Cleanup task to remove expired schedules.
    Run this daily to keep the database clean.
    """
    from .models import ExportSchedule

    cutoff_date = timezone.now().date() - timedelta(days=30)

    deleted_count = ExportSchedule.objects.filter(end_date__lt=cutoff_date).delete()[0]

    logger.info("Cleaned up %s expired schedules", deleted_count)
    return f"Deleted {deleted_count} expired schedules"

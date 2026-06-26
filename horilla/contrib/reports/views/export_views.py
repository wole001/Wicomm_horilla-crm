"""Export views for downloading report data in CSV/XLSX formats."""

# Standard library imports
import csv
from datetime import datetime

import openpyxl

# Third-party imports (Others)
import pandas as pd

# Third-party imports (Django)
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# First party imports (Horilla)
from horilla.db.models import Q
from horilla.shortcuts import get_object_or_404
from horilla.utils.decorators import method_decorator, permission_required_or_denied
from horilla.utils.translation import gettext_lazy as _
from horilla.web import HttpNotFound, HttpResponse, RefreshResponse

# Local imports
from ..models import Report
from ..views.report_detail import ReportDetailView
from .toolkit.report_helper import (
    ReportPreviewMixin,
    create_temp_report_with_preview,
    extract_display_value,
    filter_pivot_data,
)


@method_decorator(
    permission_required_or_denied("reports.change_report"), name="dispatch"
)
class ReportExportView(ReportPreviewMixin, LoginRequiredMixin, View):
    """
    Export pivot table data in various formats: Excel, CSV
    """

    def get(self, request, pk):
        """Handle GET request to export report data in Excel or CSV format."""
        try:
            report = get_object_or_404(Report, pk=pk)
        except Exception as e:
            if request.headers.get("HX-Request") == "true":
                messages.error(self.request, e)
                return RefreshResponse(request)
            raise HttpNotFound(e) from e
        export_format = request.GET.get("format", "excel")

        preview_data = self.get_preview_data(request, report)
        temp_report = create_temp_report_with_preview(report, preview_data)

        df, _context = self.get_report_data(temp_report, request)

        detail_view = ReportDetailView()
        detail_view.request = request
        detail_view.args = self.args
        detail_view.kwargs = self.kwargs
        detail_view.object = report
        detail_context = detail_view.get_context_data()

        if export_format == "excel":
            return self.export_excel(report, df, detail_context, temp_report)
        if export_format == "csv":
            return self.export_csv(report, df, detail_context, temp_report)
        # else:
        return self.export_excel(report, df, detail_context, temp_report)

    def get_report_data(self, temp_report, request):
        """Get processed report data"""
        model_class = temp_report.model_class
        queryset = model_class.objects.all()

        # Apply filters
        filters = temp_report.filters_dict
        if filters:
            query = None
            for index, (field_name, filter_data) in enumerate(filters.items()):
                if not filter_data.get("value"):
                    continue

                operator = filter_data.get("operator", "exact")
                value = filter_data.get("value")
                logic = filter_data.get("logic", "and") if index > 0 else "and"
                actual_field = filter_data.get("original_field", field_name)

                filter_kwargs = {}
                if operator == "exact":
                    filter_kwargs[f"{actual_field}"] = value
                elif operator == "icontains":
                    filter_kwargs[f"{actual_field}__icontains"] = value
                elif operator == "gt":
                    filter_kwargs[f"{actual_field}__gt"] = value
                elif operator == "lt":
                    filter_kwargs[f"{actual_field}__lt"] = value
                elif operator == "gte":
                    filter_kwargs[f"{actual_field}__gte"] = value
                elif operator == "lte":
                    filter_kwargs[f"{actual_field}__lte"] = value

                if filter_kwargs:
                    current_query = Q(**filter_kwargs)
                    if query is None:
                        query = current_query
                    elif logic == "or":
                        query |= current_query
                    else:
                        query &= current_query

            if query:
                queryset = queryset.filter(query)

        # Get fields and convert to DataFrame
        fields = []
        aggregate_columns_dict = temp_report.aggregate_columns_dict
        if not isinstance(aggregate_columns_dict, list):
            aggregate_columns_dict = (
                [aggregate_columns_dict] if aggregate_columns_dict else []
            )

        if temp_report.selected_columns_list:
            fields.extend(temp_report.selected_columns_list)
        if temp_report.row_groups_list:
            fields.extend(temp_report.row_groups_list)
        if temp_report.column_groups_list:
            fields.extend(temp_report.column_groups_list)

        for agg in aggregate_columns_dict:
            if agg.get("field"):
                fields.append(agg["field"])

        fields = list(dict.fromkeys(fields))
        data = list(queryset.values(*fields)) if fields else list(queryset.values())
        df = pd.DataFrame(data)

        # Create context for export
        context = {
            "total_count": len(data),
            "configuration_type": self.get_configuration_type(temp_report),
            "aggregate_columns_dict": aggregate_columns_dict,
        }

        return df, context

    def get_configuration_type(self, report):
        """Return configuration type string based on row and column group counts for export."""
        row_count = len(report.row_groups_list)
        col_count = len(report.column_groups_list)
        return f"{row_count}_row_{col_count}_col"

    def export_excel(self, report, df, detail_context, temp_report):
        """Export pivot table as Excel file"""
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{report.name}_pivot.xlsx"'
        )

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pivot Table"

        # Write pivot table data based on configuration type
        config_type = self.get_configuration_type(temp_report)
        self._create_excel_sheet(ws, df, detail_context, temp_report, config_type)

        # Add metadata sheet
        meta_ws = wb.create_sheet("Report Info")
        meta_ws["A1"] = "Report Name"
        meta_ws["B1"] = report.name
        meta_ws["A2"] = "Export Date"
        meta_ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        meta_ws["A3"] = "Total Records"
        meta_ws["B3"] = detail_context.get("total_count", 0)

        # Save to response
        wb.save(response)
        return response

    def _create_excel_sheet(self, ws, df, detail_context, temp_report, config_type):
        """Route to appropriate sheet creation method based on configuration"""
        if config_type == "2_row_0_col":
            self._create_hierarchical_excel_sheet(
                ws, detail_context, temp_report, "2_level"
            )
        elif config_type == "2_row_1_col":
            self._create_hierarchical_excel_sheet(
                ws, detail_context, temp_report, "2_level_with_col"
            )
        elif config_type == "3_row_0_col":
            self._create_hierarchical_excel_sheet(
                ws, detail_context, temp_report, "3_level"
            )
        else:
            # Use existing pivot table logic for 0x0, 1x0, 1x1, 1x2
            self._create_pivot_sheet(ws, df, detail_context, temp_report)

    def _create_hierarchical_excel_sheet(
        self, ws, detail_context, temp_report, hierarchy_type
    ):
        """Create Excel sheet for hierarchical data structures (2 row 0 col, 2 row 1 col, 3 row 0 col)"""

        if hierarchy_type == "3_level":
            three_level_data = detail_context.get("three_level_data", {})
            groups = three_level_data.get("groups", [])
            grand_total = three_level_data.get("grand_total", 0)
            aggregate_columns = detail_context.get("aggregate_columns", [])
            row_verbose_names = detail_context.get("row_group_verbose_names", [])

            if not groups:
                ws["A1"] = "No data available"
                return

            # Create headers - use verbose names instead of generic labels
            level1_header = (
                row_verbose_names[0]
                if row_verbose_names and len(row_verbose_names) > 0
                else "Level 1"
            )
            level2_header = (
                row_verbose_names[1]
                if row_verbose_names and len(row_verbose_names) > 1
                else "Level 2"
            )
            level3_header = (
                row_verbose_names[2]
                if row_verbose_names and len(row_verbose_names) > 2
                else "Level 3"
            )
            headers = [level1_header, level2_header, level3_header, "Count"]
            for agg in aggregate_columns:
                headers.append(agg["name"])

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=str(header))
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center")

            # Write data - Convert all display values to strings and track merging
            row_idx = 2
            for level1_group in groups:
                level1_start_row = row_idx
                for level2_group in level1_group["level2_groups"]:
                    level2_start_row = row_idx
                    for level3_item in level2_group["level3_items"]:
                        ws.cell(
                            row=row_idx,
                            column=1,
                            value=str(level1_group["level1_group_display"]),
                        )
                        ws.cell(
                            row=row_idx,
                            column=2,
                            value=str(level2_group["level2_group_display"]),
                        )
                        ws.cell(
                            row=row_idx,
                            column=3,
                            value=str(level3_item["level3_group_display"]),
                        )
                        ws.cell(row=row_idx, column=4, value=level3_item["count"])

                        col_idx = 5
                        for agg in aggregate_columns:
                            value = level3_item["aggregate_values"].get(agg["name"], 0)
                            ws.cell(row=row_idx, column=col_idx, value=value)
                            col_idx += 1
                        row_idx += 1

                    # Merge level2 cells if there are multiple level3 items
                    level2_end_row = row_idx - 1
                    if level2_end_row > level2_start_row:
                        ws.merge_cells(
                            start_row=level2_start_row,
                            start_column=2,
                            end_row=level2_end_row,
                            end_column=2,
                        )
                        ws.cell(row=level2_start_row, column=2).alignment = Alignment(
                            horizontal="center", vertical="center"
                        )

                # Merge level1 cells if there are multiple level2 groups
                level1_end_row = row_idx - 1
                if level1_end_row > level1_start_row:
                    ws.merge_cells(
                        start_row=level1_start_row,
                        start_column=1,
                        end_row=level1_end_row,
                        end_column=1,
                    )
                    ws.cell(row=level1_start_row, column=1).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

            # Add grand total
            ws.cell(row=row_idx, column=1, value="Grand Total").font = Font(bold=True)
            ws.cell(row=row_idx, column=4, value=grand_total).font = Font(bold=True)

        elif hierarchy_type in ["2_level", "2_level_with_col"]:
            hierarchical_data = detail_context.get("hierarchical_data", {})
            groups = hierarchical_data.get("groups", [])
            grand_total = hierarchical_data.get("grand_total", 0)
            pivot_columns = detail_context.get("pivot_columns", [])
            aggregate_columns = detail_context.get("aggregate_columns", [])
            row_verbose_names = detail_context.get("row_group_verbose_names", [])

            if not groups:
                ws["A1"] = "No data available"
                return

            # Create headers - use verbose names instead of generic labels
            primary_header = (
                row_verbose_names[0]
                if row_verbose_names and len(row_verbose_names) > 0
                else "Primary Group"
            )
            secondary_header = (
                row_verbose_names[1]
                if row_verbose_names and len(row_verbose_names) > 1
                else "Secondary Group"
            )
            headers = [primary_header, secondary_header]
            if hierarchy_type == "2_level_with_col":
                headers.extend(pivot_columns)
            else:
                headers.extend(
                    pivot_columns
                )  # This includes "Count" and aggregate columns

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=str(header))
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center")

            # Write data - Convert all display values to strings and track merging
            row_idx = 2
            for group in groups:
                group_start_row = row_idx
                for item in group["items"]:
                    ws.cell(
                        row=row_idx, column=1, value=str(group["primary_group_display"])
                    )
                    ws.cell(
                        row=row_idx,
                        column=2,
                        value=str(item["secondary_group_display"]),
                    )

                    col_idx = 3
                    for col_name in pivot_columns:
                        value = item["values"].get(col_name, 0)
                        ws.cell(row=row_idx, column=col_idx, value=value)
                        col_idx += 1
                    row_idx += 1

                # Merge primary group cells if there are multiple secondary items
                group_end_row = row_idx - 1
                if group_end_row > group_start_row:
                    ws.merge_cells(
                        start_row=group_start_row,
                        start_column=1,
                        end_row=group_end_row,
                        end_column=1,
                    )
                    ws.cell(row=group_start_row, column=1).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

                # Add subtotal row
                subtotal_cell = ws.cell(
                    row=row_idx,
                    column=2,
                    value=f"{str(group['primary_group_display'])} Subtotal",
                )
                subtotal_cell.font = Font(bold=True, italic=True)
                ws.cell(row=row_idx, column=3, value=group["subtotal"]).font = Font(
                    bold=True
                )
                row_idx += 1

            # Add grand total
            ws.cell(row=row_idx, column=2, value="Grand Total").font = Font(bold=True)
            ws.cell(row=row_idx, column=3, value=grand_total).font = Font(bold=True)

        # Apply borders and auto-adjust columns
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        max_row = ws.max_row
        max_col = ws.max_column

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).border = thin_border
                if row > 1 and col > 2:
                    ws.cell(row=row, column=col).alignment = Alignment(
                        horizontal="center"
                    )

        # Auto-adjust column widths
        for col_idx in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max(max_length + 2, 12)
            ws.column_dimensions[col_letter].width = min(adjusted_width, 30)

    def _create_hierarchical_csv(self, writer, detail_context, hierarchy_type):
        """Create CSV for hierarchical data structures"""

        if hierarchy_type == "3_level":
            three_level_data = detail_context.get("three_level_data", {})
            groups = three_level_data.get("groups", [])
            grand_total = three_level_data.get("grand_total", 0)
            aggregate_columns = detail_context.get("aggregate_columns", [])
            row_verbose_names = detail_context.get("row_group_verbose_names", [])

            if not groups:
                writer.writerow(["No data available"])
                return

            # Headers - use verbose names
            level1_header = (
                row_verbose_names[0]
                if row_verbose_names and len(row_verbose_names) > 0
                else "Level 1"
            )
            level2_header = (
                row_verbose_names[1]
                if row_verbose_names and len(row_verbose_names) > 1
                else "Level 2"
            )
            level3_header = (
                row_verbose_names[2]
                if row_verbose_names and len(row_verbose_names) > 2
                else "Level 3"
            )
            headers = [level1_header, level2_header, level3_header, "Count"]
            for agg in aggregate_columns:
                headers.append(agg["name"])
            writer.writerow(headers)

            # Data rows
            for level1_group in groups:
                for level2_group in level1_group["level2_groups"]:
                    for level3_item in level2_group["level3_items"]:
                        row = [
                            level1_group["level1_group_display"],
                            level2_group["level2_group_display"],
                            level3_item["level3_group_display"],
                            level3_item["count"],
                        ]
                        for agg in aggregate_columns:
                            row.append(
                                level3_item["aggregate_values"].get(agg["name"], 0)
                            )
                        writer.writerow(row)

            # Grand total
            total_row = ["Grand Total", "", "", grand_total]
            writer.writerow(total_row)

        elif hierarchy_type in ["2_level", "2_level_with_col"]:
            hierarchical_data = detail_context.get("hierarchical_data", {})
            groups = hierarchical_data.get("groups", [])
            grand_total = hierarchical_data.get("grand_total", 0)
            pivot_columns = detail_context.get("pivot_columns", [])
            row_verbose_names = detail_context.get("row_group_verbose_names", [])

            if not groups:
                writer.writerow(["No data available"])
                return

            # Headers - use verbose names
            primary_header = (
                row_verbose_names[0]
                if row_verbose_names and len(row_verbose_names) > 0
                else "Primary Group"
            )
            secondary_header = (
                row_verbose_names[1]
                if row_verbose_names and len(row_verbose_names) > 1
                else "Secondary Group"
            )
            headers = [primary_header, secondary_header]
            headers.extend(pivot_columns)
            writer.writerow(headers)

            # Data rows
            for group in groups:
                for item in group["items"]:
                    row = [
                        group["primary_group_display"],
                        item["secondary_group_display"],
                    ]
                    for col_name in pivot_columns:
                        row.append(item["values"].get(col_name, 0))
                    writer.writerow(row)

                # Subtotal row
                subtotal_row = [
                    "",
                    f"{group['primary_group_display']} Subtotal",
                    group["subtotal"],
                ]
                writer.writerow(subtotal_row)

            # Grand total
            writer.writerow(["", "Grand Total", grand_total])

    def _create_pivot_sheet(self, ws, df, detail_context, temp_report):
        """Create pivot table sheet that matches the web detail view"""
        pivot_table = detail_context.get("pivot_table", {})
        pivot_index = detail_context.get("pivot_index", [])
        pivot_columns = detail_context.get("pivot_columns", [])
        row_verbose_names = detail_context.get("row_group_verbose_names", [])

        # FILTER OUT ID-ONLY KEYS
        pivot_table, pivot_index, pivot_columns = filter_pivot_data(
            pivot_table, pivot_index, pivot_columns
        )

        if not pivot_table:
            # Handle 0x0 configuration (simple aggregate)
            simple_aggregate = detail_context.get("simple_aggregate", {})
            aggregate_columns = detail_context.get("aggregate_columns", [])

            if simple_aggregate or aggregate_columns:
                ws["A1"] = "Metric"
                ws["B1"] = "Value"
                ws["A1"].font = Font(bold=True)
                ws["B1"].font = Font(bold=True)
                ws["A1"].fill = PatternFill(
                    start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"
                )
                ws["B1"].fill = PatternFill(
                    start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"
                )

                row_idx = 2
                if aggregate_columns:
                    for agg in aggregate_columns:
                        ws.cell(row=row_idx, column=1, value=agg["name"])
                        ws.cell(row=row_idx, column=2, value=agg["value"])
                        row_idx += 1
                elif simple_aggregate:
                    metric_name = f"{simple_aggregate['function'].title()} of {simple_aggregate['field']}"
                    ws.cell(row=row_idx, column=1, value=metric_name)
                    ws.cell(row=row_idx, column=2, value=simple_aggregate["value"])
            else:
                ws["A1"] = "No pivot table data available"
            return

        if not pivot_index:
            ws["A1"] = "No data available"
            return

        row_header = row_verbose_names[0] if row_verbose_names else "Row Group"

        has_hierarchical_columns = any(
            col.count("|") == 3 and "||" in col for col in pivot_columns
        )

        has_simple_groups = len(pivot_columns) > 1 and not has_hierarchical_columns

        if has_hierarchical_columns:
            self._create_hierarchical_column_sheet(
                ws, pivot_table, pivot_index, pivot_columns, row_header
            )
        elif has_simple_groups:
            self._create_simple_header_sheet(
                ws, pivot_table, pivot_index, pivot_columns, row_header
            )
        else:
            self._create_simple_header_sheet(
                ws, pivot_table, pivot_index, pivot_columns, row_header
            )

    def _create_hierarchical_column_sheet(
        self, ws, pivot_table, pivot_index, pivot_columns, row_header
    ):
        """Create sheet with hierarchical column headers (for 1 row × 2 col)"""
        # Parse hierarchical columns: "Group1Display||Group1ID|Group2Display||Group2ID"
        ws.cell(row=1, column=1, value=row_header)
        ws.cell(row=2, column=1, value=row_header)
        ws.merge_cells("A1:A2")
        ws["A1"].font = Font(bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        current_col = 2
        group_starts = {}
        current_group = None
        group_start_col = 2

        # Parse column structure
        for col_name in pivot_columns:
            if "|" in col_name and "||" in col_name:
                # Split by single | first to separate the two groups
                main_parts = col_name.split("|")
                if main_parts[1] == "":  # This means we have || in the string
                    # Reconstruct: parts[0] is Display1, parts[2] is ID1, parts[3] is Display2, parts[5] is ID2
                    if len(main_parts) >= 4:
                        group1_display = self.extract_display_value(
                            main_parts[0] + "||" + main_parts[2]
                        )
                        if len(main_parts) >= 6:
                            group2_display = extract_display_value(
                                main_parts[3] + "||" + main_parts[5]
                            )
                        else:
                            group2_display = extract_display_value(main_parts[3])
                    else:
                        # Fallback
                        group1_display = extract_display_value(col_name)
                        group2_display = ""
                else:
                    # Try to find || positions
                    first_double_pipe = col_name.find("||")
                    if first_double_pipe != -1:
                        # Find the single | after the first ||
                        search_start = first_double_pipe + 2
                        next_single_pipe = col_name.find("|", search_start)
                        while (
                            next_single_pipe != -1
                            and next_single_pipe < len(col_name) - 1
                            and col_name[next_single_pipe + 1] == "|"
                        ):
                            next_single_pipe = col_name.find("|", next_single_pipe + 2)

                        if next_single_pipe != -1:
                            # Split at this position
                            group1_composite = col_name[:next_single_pipe]
                            group2_composite = col_name[next_single_pipe + 1 :]
                            group1_display = extract_display_value(group1_composite)
                            group2_display = extract_display_value(group2_composite)
                        else:
                            group1_display = extract_display_value(col_name)
                            group2_display = ""
                    else:
                        group1_display = extract_display_value(col_name)
                        group2_display = ""

                if current_group != group1_display:
                    if current_group is not None:
                        group_starts[current_group] = (group_start_col, current_col - 1)
                    current_group = group1_display
                    group_start_col = current_col

                # Set headers
                ws.cell(row=1, column=current_col, value=group1_display).font = Font(
                    bold=True
                )
                ws.cell(row=2, column=current_col, value=group2_display).font = Font(
                    bold=True
                )
            else:
                # Aggregate column (no || present)
                display_name = extract_display_value(col_name)
                ws.cell(row=1, column=current_col, value=display_name).font = Font(
                    bold=True
                )
                ws.cell(row=2, column=current_col, value="").font = Font(bold=True)

            current_col += 1

        # Record last group
        if current_group is not None:
            group_starts[current_group] = (group_start_col, current_col - 1)

        # Merge group headers
        for _group_name, (start_col, end_col) in group_starts.items():
            if start_col < end_col:
                start_cell = openpyxl.utils.get_column_letter(start_col)
                end_cell = openpyxl.utils.get_column_letter(end_col)
                ws.merge_cells(f"{start_cell}1:{end_cell}1")
                ws[f"{start_cell}1"].alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, row_key in enumerate(pivot_index, 3):
            display_value = extract_display_value(row_key)
            ws.cell(row=row_idx, column=1, value=display_value)

            for col_idx, col_name in enumerate(pivot_columns, 2):
                value = pivot_table.get(row_key, {}).get(col_name, 0)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Add total row
        total_row = len(pivot_index) + 3
        ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)

        for col_idx, col_name in enumerate(pivot_columns, 2):
            total_value = sum(
                pivot_table.get(row_key, {}).get(col_name, 0) for row_key in pivot_index
            )
            ws.cell(row=total_row, column=col_idx, value=total_value).font = Font(
                bold=True
            )

        # Apply styling
        self._apply_excel_styling(ws, total_row, len(pivot_columns) + 1, header_rows=2)

    def _create_simple_header_sheet(
        self, ws, pivot_table, pivot_index, pivot_columns, row_header
    ):
        """Create pivot sheet with single-row header"""
        # Single header row
        ws.cell(row=1, column=1, value=row_header)
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # Column headers
        for col_idx, col_name in enumerate(pivot_columns, 2):
            display_name = extract_display_value(col_name)
            cell = ws.cell(row=1, column=col_idx, value=display_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row_key in enumerate(pivot_index, 2):
            display_value = extract_display_value(row_key)
            ws.cell(row=row_idx, column=1, value=display_value)

            for col_idx, col_name in enumerate(pivot_columns, 2):
                value = pivot_table.get(row_key, {}).get(col_name, 0)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Total row
        total_row = len(pivot_index) + 2
        ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)

        for col_idx, col_name in enumerate(pivot_columns, 2):
            total_value = sum(
                pivot_table.get(row_key, {}).get(col_name, 0) for row_key in pivot_index
            )
            ws.cell(row=total_row, column=col_idx, value=total_value).font = Font(
                bold=True
            )

        # Apply styling
        self._apply_excel_styling(ws, total_row, len(pivot_columns) + 1, header_rows=1)

    def _apply_excel_styling(self, ws, max_row, max_col, header_rows=1):
        """Apply consistent styling to Excel sheets"""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Apply borders and alignment
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

                # Center align data cells
                if row > header_rows and col > 1:
                    cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for col_idx in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_length = 0

            for row_idx in range(1, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = (
                max(max_length + 2, 12) if col_idx == 1 else max(max_length + 2, 10)
            )
            ws.column_dimensions[col_letter].width = min(adjusted_width, 30)

        # Add header background colors
        for col_idx in range(1, max_col + 1):
            for header_row in range(1, header_rows + 1):
                cell = ws.cell(row=header_row, column=col_idx)
                if header_row == 1:
                    cell.fill = PatternFill(
                        start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"
                    )
                else:
                    cell.fill = PatternFill(
                        start_color="E8F4FD", end_color="E8F4FD", fill_type="solid"
                    )

    def export_csv(self, report, df, detail_context, temp_report):
        """Export pivot table as CSV"""
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="{report.name}_pivot.csv"'
        )

        writer = csv.writer(response)
        config_type = self.get_configuration_type(temp_report)

        if config_type == "2_row_0_col":
            self._create_hierarchical_csv(writer, detail_context, "2_level")
        elif config_type == "2_row_1_col":
            self._create_hierarchical_csv(writer, detail_context, "2_level_with_col")
        elif config_type == "3_row_0_col":
            self._create_hierarchical_csv(writer, detail_context, "3_level")
        else:
            self._create_pivot_csv(writer, detail_context, temp_report)

        return response

    def _create_pivot_csv(self, writer, detail_context, temp_report):
        """Create CSV for pivot table data"""
        pivot_table = detail_context.get("pivot_table", {})
        pivot_index = detail_context.get("pivot_index", [])
        pivot_columns = detail_context.get("pivot_columns", [])
        row_verbose_names = detail_context.get("row_group_verbose_names", [])

        # FILTER OUT ID-ONLY KEYS
        pivot_table, pivot_index, pivot_columns = filter_pivot_data(
            pivot_table, pivot_index, pivot_columns
        )

        if not pivot_table:
            # Handle 0x0 configuration
            simple_aggregate = detail_context.get("simple_aggregate", {})
            aggregate_columns = detail_context.get("aggregate_columns", [])

            if simple_aggregate or aggregate_columns:
                writer.writerow(["Metric", "Value"])
                if aggregate_columns:
                    for agg in aggregate_columns:
                        writer.writerow([agg["name"], agg["value"]])
                elif simple_aggregate:
                    metric_name = f"{simple_aggregate['function'].title()} of {simple_aggregate['field']}"
                    writer.writerow([metric_name, simple_aggregate["value"]])
            else:
                writer.writerow(["No pivot table data available"])
            return

        if not pivot_index:
            writer.writerow(["No data available"])
            return

        row_header = row_verbose_names[0] if row_verbose_names else "Row Group"

        # Check for hierarchical columns (1 row × 2 col)
        has_hierarchical_columns = any(
            col.count("|") == 3 and "||" in col for col in pivot_columns
        )

        if has_hierarchical_columns:
            # Two-row header for hierarchical columns
            group_header = [row_header]
            column_header = [row_header]

            for col_name in pivot_columns:
                if "|" in col_name and "||" in col_name:
                    # Find the pattern: "Display1||ID1|Display2||ID2"
                    first_double_pipe = col_name.find("||")
                    if first_double_pipe != -1:
                        search_start = first_double_pipe + 2
                        next_single_pipe = col_name.find("|", search_start)
                        while (
                            next_single_pipe != -1
                            and next_single_pipe < len(col_name) - 1
                            and col_name[next_single_pipe + 1] == "|"
                        ):
                            next_single_pipe = col_name.find("|", next_single_pipe + 2)

                        if next_single_pipe != -1:
                            group1_composite = col_name[:next_single_pipe]
                            group2_composite = col_name[next_single_pipe + 1 :]
                            group1_display = extract_display_value(group1_composite)
                            group2_display = extract_display_value(group2_composite)
                        else:
                            group1_display = extract_display_value(col_name)
                            group2_display = ""
                    else:
                        group1_display = extract_display_value(col_name)
                        group2_display = ""

                    group_header.append(group1_display)
                    column_header.append(group2_display)
                else:
                    display_name = extract_display_value(col_name)
                    group_header.append(display_name)
                    column_header.append("")

            writer.writerow(group_header)
            writer.writerow(column_header)
        else:
            # Single-row header
            header = [row_header]
            for col_name in pivot_columns:
                display_name = extract_display_value(col_name)
                header.append(display_name)
            writer.writerow(header)

        # Write data rows
        for row_key in pivot_index:
            display_value = extract_display_value(row_key)
            row = [display_value]
            for col_name in pivot_columns:
                value = pivot_table.get(row_key, {}).get(col_name, 0)
                row.append(value)
            writer.writerow(row)

        # Add totals row
        total_row = ["Total"]
        for col_name in pivot_columns:
            total_value = sum(
                pivot_table.get(row_key, {}).get(col_name, 0) for row_key in pivot_index
            )
            total_row.append(total_value)
        writer.writerow(total_row)

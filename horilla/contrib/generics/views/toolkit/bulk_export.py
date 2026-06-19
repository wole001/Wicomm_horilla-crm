"""
Bulk export mixin extracted from HorillaListView to keep the main
list view class smaller and focused.
"""

# Standard library imports
import csv
import json
import logging
import re
from datetime import date, datetime
from io import BytesIO

# Third-party imports (Django)
from django.db.models.fields.related import ManyToManyField
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from horilla.contrib.core.utils import sanitize_export_value
from horilla.db.models import ForeignKey

# First party imports (Horilla)
from horilla.utils import timezone
from horilla.web import HttpResponse

logger = logging.getLogger(__name__)


class HorillaBulkExportMixin:
    """
    Mixin that encapsulates bulk export logic (CSV/XLSX/PDF) so that
    list views can stay focused on querying and rendering.
    """

    def handle_bulk_export_post(self, record_ids, columns, export_format):
        """
        Entry point used from HorillaListView.post.

        - Safely decodes record_ids from JSON string
        - Delegates to handle_export for actual file generation
        - Returns HttpResponse or None if export is not requested
        """
        if not (record_ids and export_format):
            return None

        try:
            record_ids_list = json.loads(record_ids)
        except json.JSONDecodeError:
            return HttpResponse("Invalid JSON data for record_ids", status=400)

        return HorillaBulkExportMixin.handle_export(
            self, record_ids_list, columns, export_format
        )

    def handle_export(self, record_ids, columns, export_format):
        """
        Handle the export of data in the specified format.

        NOTE: This logic was moved out of HorillaListView for clarity but
        kept functionally identical.
        """

        try:
            queryset = self.model.objects.filter(id__in=record_ids)
            model_fields = [
                (str(field.verbose_name), field.name, field)
                for field in self.model._meta.fields
            ]
            property_labels = getattr(self.model, "PROPERTY_LABELS", None)
            # Only process properties if PROPERTY_LABELS is explicitly defined on the model
            if property_labels:
                for name, label in property_labels.items():
                    member = getattr(self.model, name, None)
                    if member is None:
                        continue
                    if isinstance(member, property) or callable(member):
                        label_key = (
                            name.replace("get_", "", 1)
                            if name.startswith("get_")
                            else name
                        )
                        if label_key in ["histories", "full_histories"]:
                            continue
                        model_fields.append((str(label), name, None))

            for field in self.model._meta.fields:
                if field.choices:
                    method_name = f"get_{field.name}_display"
                    if hasattr(self.model, method_name):
                        model_fields.append(
                            (str(field.verbose_name), method_name, "method")
                        )

            # Get table columns from _get_columns
            table_columns = self._get_columns()
            # Filter out histories and full_histories columns from export
            exclude_from_export = ["histories", "full_histories"]
            table_columns = [
                col for col in table_columns if col[1] not in exclude_from_export
            ]
            table_column_names = [
                col[1] for col in table_columns
            ]  # Field names of table columns

            # Use selected columns if provided, otherwise use table columns
            if columns:
                # Filter out histories and full_histories from selected columns
                columns = [col for col in columns if col not in exclude_from_export]
                column_headers = [
                    field[0] for field in model_fields if field[1] in columns
                ]
                selected_fields = [
                    field for field in model_fields if field[1] in columns
                ]
            else:
                # Use table columns instead of all model fields
                column_headers = [col[0] for col in table_columns]
                selected_fields = [
                    field for field in model_fields if field[1] in table_column_names
                ]

                # If no table columns are defined, log error and return
                if not table_columns:
                    return HttpResponse(
                        "No table columns defined for export", status=400
                    )

            data = []
            for obj in queryset:
                row = []
                for _verbose_name, field_name, field in selected_fields:
                    try:
                        # Prefer display value for choice fields (e.g. role, country)
                        # and CountryField (e.g. country) so export shows label not key
                        display_method_name = f"get_{field_name}_display"
                        if hasattr(obj, display_method_name):
                            display_method = getattr(obj, display_method_name)
                            if callable(display_method):
                                value = display_method()
                            else:
                                value = getattr(obj, field_name, "")
                        else:
                            value = getattr(obj, field_name, "")
                        if field == "method" or callable(value):
                            value = value() if callable(value) else value
                            if (
                                field is None
                                and value
                                and isinstance(value, str)
                                and "<" in value
                            ):
                                value = re.sub(r"<[^>]+>", "", value).strip()
                        elif field is None:  # This is a @property
                            if callable(value):
                                value = value()
                            if value and isinstance(value, str) and "<" in value:
                                value = re.sub(r"<[^>]+>", "", value).strip()
                        elif isinstance(field, ForeignKey):
                            value = (
                                str(getattr(value, "username", value)) if value else ""
                            )
                        elif isinstance(field, ManyToManyField):
                            # Handle ManyToManyField (e.g., tags)
                            value = (
                                ", ".join(str(item) for item in value.all())
                                if value
                                else ""
                            )
                        elif callable(value):
                            value = value()
                        # Format date/datetime with user's chosen format and timezone (not UTC)
                        user = getattr(self.request, "user", None)
                        if user and value is not None:
                            if isinstance(value, datetime):
                                if getattr(user, "time_zone", None):
                                    try:
                                        from zoneinfo import ZoneInfo

                                        user_tz = ZoneInfo(user.time_zone)
                                        if timezone.is_naive(value):
                                            value = timezone.make_aware(
                                                value, timezone.get_default_timezone()
                                            )
                                        value = value.astimezone(user_tz)
                                    except Exception:
                                        pass
                                fmt = (
                                    getattr(user, "date_time_format", None)
                                    or "%Y-%m-%d %H:%M:%S"
                                )
                                try:
                                    value = value.strftime(fmt)
                                except Exception:
                                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                            elif isinstance(value, date) and not isinstance(
                                value, datetime
                            ):
                                fmt = getattr(user, "date_format", None) or "%Y-%m-%d"
                                try:
                                    value = value.strftime(fmt)
                                except Exception:
                                    value = value.strftime("%Y-%m-%d")
                        row.append(str(value) if value is not None else "")
                    except Exception as e:
                        logger.error(
                            "Error retrieving field %s: %s", field_name, str(e)
                        )
                        row.append("")  # Fallback to empty string
                data.append(row)

            model_verbose_name = self.model._meta.verbose_name_plural.lower().replace(
                " ", "_"
            )
            document_title = f"Exported {self.model._meta.verbose_name_plural}"
            if export_format == "csv":
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    f'attachment; filename="exported_{model_verbose_name}.csv"'
                )
                writer = csv.writer(response)
                writer.writerow(column_headers)
                for row in data:
                    sanitized_row = [sanitize_export_value(cell) for cell in row]
                    writer.writerow(sanitized_row)
                return response

            if export_format == "xlsx":
                wb = Workbook()
                ws = wb.active

                # Append the header row
                ws.append([str(header) for header in column_headers])

                # Style the header row
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal="center")
                header_fill = PatternFill(
                    start_color="eafb5b", end_color="eafb5b", fill_type="solid"
                )  # Light gray background

                for cell in ws[1]:  # First row (headers)
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.fill = header_fill

                # Set column widths (adjust the width as needed)
                for col in ws.columns:
                    column_letter = col[0].column_letter
                    ws.column_dimensions[column_letter].width = 25

                # Append the data rows
                for row in data:
                    sanitized_row = [
                        sanitize_export_value(str(cell) if cell is not None else "")
                        for cell in row
                    ]
                    ws.append(sanitized_row)

                for idx, row in enumerate(ws.rows, 1):
                    ws.row_dimensions[idx].height = 15

                response = HttpResponse(
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = (
                    f'attachment; filename="exported_{model_verbose_name}.xlsx"'
                )
                buffer = BytesIO()
                wb.save(buffer)
                response.write(buffer.getvalue())
                buffer.close()
                return response

            if export_format == "pdf":
                buffer = BytesIO()
                # Use landscape orientation for better width
                page_size = (letter[1], letter[0])  # 792 x 612 points (landscape)
                width, height = page_size

                c = canvas.Canvas(buffer, pagesize=page_size)

                # Set PDF metadata title
                c.setTitle(document_title)

                # Set fonts and adjust font size for readability
                title_font_size = 18
                header_font_size = 12
                data_font_size = 10

                # Table settings
                start_x = 50
                start_y = height - 100
                min_col_width = 120
                # padding = 8
                max_rows_per_page = 7
                max_cols_per_page = 6
                extra_row_spacing = 10

                # Helper function to wrap text
                def wrap_text(text, max_chars):
                    text = str(text) if text is not None else ""
                    if len(text) <= max_chars:
                        return [text] if text else [""]
                    words = text.split()
                    lines = []
                    current_line = ""
                    for word in words:
                        if len(current_line) + len(word) + 1 <= max_chars:
                            current_line += word + " "
                        else:
                            lines.append(current_line.strip())
                            current_line = word + " "
                    if current_line:
                        lines.append(current_line.strip())
                    return lines or [""]

                # Title
                c.setFont("Helvetica-Bold", title_font_size)
                c.drawCentredString(width / 2, height - 50, document_title)

                # Draw table content with pagination for both rows and columns
                col_start_index = 0
                total_cols = len(column_headers)
                pages_created = False

                while col_start_index < total_cols:
                    current_col_headers = column_headers[
                        col_start_index : col_start_index + max_cols_per_page
                    ]

                    if pages_created:
                        c.showPage()
                        c.setTitle(document_title)

                    # Draw headers
                    c.setFont("Helvetica-Bold", header_font_size)
                    x = start_x
                    y = start_y
                    col_widths = [min_col_width] * len(current_col_headers)
                    for i, header in enumerate(current_col_headers):
                        wrapped_header = wrap_text(header, 18)
                        line_height = header_font_size + 2
                        # total_header_height = len(wrapped_header) * line_height
                        for j, header_line in enumerate(wrapped_header):
                            c.drawString(
                                x,
                                y - (j * line_height),
                                header_line[: 50 if i == 0 else 30],
                            )
                        max_width_for_col = min_col_width
                        col_widths[i] = max(min_col_width, max_width_for_col)
                        x += col_widths[i]
                    y -= (header_font_size + 2) * max(
                        len(wrap_text(h, 18)) for h in current_col_headers
                    ) + extra_row_spacing

                    # Draw data rows
                    c.setFont("Helvetica", data_font_size)
                    rows_on_page = 0
                    for row_data in data:
                        if rows_on_page >= max_rows_per_page:
                            c.showPage()
                            c.setTitle(document_title)
                            c.setFont("Helvetica-Bold", header_font_size)
                            x = start_x
                            y = start_y
                            for i, header in enumerate(current_col_headers):
                                wrapped_header = wrap_text(header, 18)
                                line_height = header_font_size + 2
                                # total_header_height = len(wrapped_header) * line_height
                                for j, header_line in enumerate(wrapped_header):
                                    c.drawString(
                                        x,
                                        y - (j * line_height),
                                        header_line[: 50 if i == 0 else 30],
                                    )
                                x += col_widths[i]
                            y -= (header_font_size + 2) * max(
                                len(wrap_text(h, 18)) for h in current_col_headers
                            ) + extra_row_spacing
                            c.setFont("Helvetica", data_font_size)
                            rows_on_page = 0

                        x = start_x
                        row_height = 0
                        for i, col_index in enumerate(
                            range(
                                col_start_index,
                                min(col_start_index + max_cols_per_page, total_cols),
                            )
                        ):
                            value = row_data[col_index]
                            max_chars = 20 if i == 0 else 15
                            wrapped_text = wrap_text(value, max_chars)
                            line_height = data_font_size + 2
                            total_height = len(wrapped_text) * line_height
                            for j, line in enumerate(wrapped_text):
                                c.drawString(
                                    x,
                                    y - (j * line_height),
                                    line[: 60 if i == 0 else 40],
                                )
                            x += col_widths[i]
                            row_height = max(row_height, total_height)
                        y -= row_height + extra_row_spacing
                        rows_on_page += 1

                    pages_created = True
                    col_start_index += max_cols_per_page

                c.showPage()
                c.save()

                buffer.seek(0)
                response = HttpResponse(content_type="application/pdf")
                response["Content-Disposition"] = (
                    f'attachment; filename="exported_{model_verbose_name}.pdf"'
                )
                response.write(buffer.getvalue())
                buffer.close()
                return response

            return HttpResponse("Invalid export format specified", status=400)

        except Exception as e:
            logger.error("Error during export: %s", str(e))
            return HttpResponse("An error occurred during export", status=500)
